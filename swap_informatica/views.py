import openpyxl
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment, NamedStyle
from openpyxl.formatting.rule import CellIsRule
import json
from datetime import timedelta, datetime, date

from django.db import IntegrityError, transaction
from django.shortcuts import render, redirect, get_object_or_404
from django.utils.timezone import now
from django.db.models import Q
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
from django.utils import timezone
from django.contrib.auth.decorators import login_required, user_passes_test
from django.utils.dateparse import parse_date
from django.http import HttpResponse
from django.db.models.functions import ExtractMonth
from django.views.decorators.http import require_POST
from django.views.generic import TemplateView
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.contrib import messages
from django.core.paginator import Paginator

from .utils import crear_mantenimiento_desde_calendario
from .forms import *
from swap_home.models import *
from swap_informatica.models import *
from swap_porteria.models import *
from swap_serviciotecnico.models import *
from swap_tecnico.models import *
from swap_tthh.models import *



# --- Proteger de otros Usuarios Logeados ---
def is_swap_informatica(user):
    return user.groups.filter(name='Informatica').exists()


# ==================== VISTA PRINCIPAL ====================
class InformaticaDashboardView(LoginRequiredMixin, UserPassesTestMixin, TemplateView):
    template_name = 'informatica.html'
    
    def test_func(self):
        return is_swap_informatica(self.request.user)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Aquí puedes agregar métricas, contadores, etc si quieres
        return context

# @login_required
# @user_passes_test(is_swap_informatica)
# def actualizar_estado(request, tarea_id):
#     if request.method == "POST":
#         tarea = get_object_or_404(Tareas, pk=tarea_id)
#         # Obtener el nuevo estado desde el formulario
#         nuevo_estado = request.POST.get("estado")
#         # Buscar el estado en la base de datos
#         estado_obj = get_object_or_404(Tareaestado, te_estado=nuevo_estado)
#         # Actualizar la tarea con el nuevo estado
#         tarea.tarea_te = estado_obj
#         tarea.save()
#     return redirect("informatica")  # Redirige a la misma página después de actualizar
# ==================== VISTA PRINCIPAL ====================



# ==================== BACKUPS ====================
@login_required
@user_passes_test(is_swap_informatica)
def backups(request):
    bk = Backups.objects.filter(b_estado='activo')
    bkechos = Backupshechos.objects.all()
    bkdia = Backupsdia.objects.all()
    bkestado = Backupsestado.objects.all()

    bkproceso = Backupsproceso.objects.exclude(bp_be__be_estado="Inactivo❌")

    form = BackupForm()
    form2 = BackupsProcesoForm()
    funcionarios = Funcionarios.objects.filter(fun_estado='activo')

    fce = Funcionarioconequipo.objects.filter(fce_estado='activo')
    form3 = fceForm()

    return render(request, 'backups.html', {
        'bk':bk,
        'bkechos': bkechos,
        'bkdia': bkdia,
        'bkestado':bkestado,
        'bkproceso':bkproceso,
        'form': form,
        'form2':form2,
        'funcionarios':funcionarios,
        'fce':fce,
        'form3': form3,
    })



def week_start_of(date_obj: datetime.date) -> datetime.date:
    """Lunes ISO de la semana para una fecha."""
    return date_obj - datetime.timedelta(days=date_obj.weekday())


@login_required
@user_passes_test(is_swap_informatica)
def backups_semanales(request):
    """
    ÚNICA VISTA (un endpoint):
    - GET: muestra programación semanal y checklist de guardado
    - POST: acciones por 'action':
        * update_estado
        * add_programado
        * guardar_realizados

    Reglas:
    - Reset semanal persistente: primera visita de una nueva semana hace reset a Pendiente⏳
    - Guardar hechos: 1 registro por (bp, week_start) -> NO duplica en la semana
    - bh_fecha = fecha real del día que guardas
    - Estado Finalizado✅ se mantiene durante la semana (hasta el reset)
    """

    hoy = timezone.localdate()
    week_start = week_start_of(hoy)
    week_end = week_start + datetime.timedelta(days=6)

    # =========================
    # 1) RESET SEMANAL PERSISTENTE
    # =========================
    with transaction.atomic():
        control = BackupsSemanaControl.objects.select_for_update().first()

        # Primera vez: crea control con la semana actual, sin resetear nada
        if not control:
            BackupsSemanaControl.objects.create(last_week_start=week_start)
        else:
            # Si cambió la semana, resetea
            if control.last_week_start < week_start:
                estado_pendiente = get_object_or_404(Backupsestado, be_estado="Pendiente⏳")

                # Resetea activos + también los que estén NULL
                Backupsproceso.objects.filter(bp_be__isnull=True).update(bp_be=estado_pendiente)
                Backupsproceso.objects.exclude(bp_be__be_estado__icontains="inactivo").update(bp_be=estado_pendiente)

                control.last_week_start = week_start
                control.save(update_fields=["last_week_start"])

                messages.success(request, "🔄 Nueva semana iniciada: todos los backups vuelven a Pendiente⏳.")

    # =========================
    # 2) POST ACTIONS
    # =========================
    if request.method == "POST":
        action = (request.POST.get("action") or "").strip()

        # ---- A) Cambiar estado ----
        if action == "update_estado":
            bp_id = request.POST.get("bp_id")
            nuevo_estado = request.POST.get("bkestado")

            if not bp_id or not nuevo_estado:
                messages.error(request, "❌ Faltan datos para actualizar el estado.")
                return redirect("backups-semanales")

            bp = get_object_or_404(Backupsproceso, pk=bp_id)
            estado_obj = get_object_or_404(Backupsestado, be_estado=nuevo_estado)

            bp.bp_be = estado_obj
            bp.save(update_fields=["bp_be"])

            messages.success(request, f"✅ Estado actualizado a {nuevo_estado}.")
            return redirect("backups-semanales")

        # ---- B) Agregar programado ----
        if action == "add_programado":
            form = BackupForm(request.POST)
            if form.is_valid():
                form.save()
                messages.success(request, "✅ Backup semanal agregado correctamente.")
            else:
                messages.error(request, "❌ Error al agregar backup semanal.")
            return redirect("backups-semanales")

        # ---- C) Guardar realizados (NO duplica por semana) ----
        if action == "guardar_realizados":
            bp_ids = request.POST.getlist("backups_realizados")

            if not bp_ids:
                messages.warning(request, "⚠️ No seleccionaste ningún backup.")
                return redirect("backups-semanales")

            estado_finalizado = get_object_or_404(Backupsestado, be_estado="Finalizado✅")

            guardados = 0
            ya_existian = 0

            with transaction.atomic():
                # Carga en bulk para evitar N queries
                procesos = (
                    Backupsproceso.objects
                    .select_for_update()
                    .filter(bp_id__in=bp_ids)
                )
                procesos_por_id = {p.bp_id: p for p in procesos}

                for raw_id in bp_ids:
                    try:
                        bp_id_int = int(raw_id)
                    except (TypeError, ValueError):
                        continue

                    bp = procesos_por_id.get(bp_id_int)
                    if not bp:
                        continue

                    # Inserta y si ya existe, no duplica (UniqueConstraint)
                    try:
                        Backupshechos.objects.create(
                            bh_bp=bp,
                            bh_fecha=hoy,
                            bh_week_start=week_start,
                        )
                        guardados += 1
                    except IntegrityError:
                        ya_existian += 1

                    # Mantener finalizado durante la semana
                    bp.bp_be = estado_finalizado
                    bp.save(update_fields=["bp_be"])

            if guardados:
                messages.success(request, f"✅ {guardados} guardado(s) como hechos el {hoy.strftime('%d/%m/%Y')}.")
            if ya_existian:
                messages.info(request, f"📌 {ya_existian} ya estaban guardados esta semana (no se duplicaron).")

            return redirect("backups-semanales")

        messages.error(request, "❌ Acción no válida.")
        return redirect("backups-semanales")

    # =========================
    # 3) GET DATA (render)
    # =========================
    bkdia = Backupsdia.objects.all().order_by("bd_id")
    bkestado = Backupsestado.objects.exclude(be_estado__icontains="inactivo")

    bkproceso_qs = (
        Backupsproceso.objects
        .select_related("bp_bd", "bp_b", "bp_be", "bp_b__b_fce")
        .exclude(bp_be__be_estado__icontains="inactivo")
        .order_by("bp_bd__bd_id", "bp_b__b_nombre")
    )

    # Hechos de hoy
    hechos_hoy = (
        Backupshechos.objects
        .filter(bh_fecha=hoy)
        .select_related("bh_bp", "bh_bp__bp_bd", "bh_bp__bp_b", "bh_bp__bp_b__b_fce")
        .order_by("bh_bp__bp_bd__bd_id")
    )

    # Hechos de la semana
    hechos_semana_ids = set(
        Backupshechos.objects
        .filter(bh_week_start=week_start)
        .values_list("bh_bp_id", flat=True)
    )

    # Agrupar backups por día (para no hacer doble loop en template)
    proceso_por_dia = {d.bd_dia: [] for d in bkdia}
    for p in bkproceso_qs:
        proceso_por_dia.setdefault(p.bp_bd.bd_dia, []).append(p)

    context = {
        "hoy": hoy,
        "week_start": week_start,
        "week_end": week_end,

        "bkdia": bkdia,
        "bkestado": bkestado,

        "bkproceso": bkproceso_qs,                 # por si quieres métricas
        "proceso_por_dia": proceso_por_dia,        # <-- usado por el template

        "hechos_hoy": hechos_hoy,
        "hechos_semana_ids": hechos_semana_ids,

        "form": BackupForm(),
    }
    return render(request, "backups/backups_semanales.html", context)



# ACTUALIZAR ESTADO BACKUP SEMANAL
# @login_required
# @user_passes_test(is_swap_informatica)
# def actualizar_estado_bk(request, bp_id):
#     if request.method == "POST":
#         bp = get_object_or_404(Backupsproceso, pk=bp_id)

#         nuevo_estado = request.POST.get("bkestado")

#         estado_obj = get_object_or_404(Backupsestado, be_estado=nuevo_estado)

#         bp.bp_be = estado_obj
#         bp.save()
#     return redirect("backups")

# #   AÑADIR BACKUPS SEMANALES
# @login_required
# @user_passes_test(is_swap_informatica)
# def add_backup(request):
#     if request.method == "POST":
#         form = BackupForm(request.POST)
#         if form.is_valid():
#             form.save()
#             return redirect("backups")
#     return redirect("backups")



# # DARLE/ELIMINAR UN BACKUP A UN FUNCIONARIO CON EQUIPO

@login_required
@user_passes_test(is_swap_informatica)
def bk_asignacion(request):
    """
    Gestión de Backups:
    - Listado de backups activos (paginado)
    - Listado de funcionarios con equipo SIN backup (paginado)
    - Form para asignar backup
    - Form para desactivar backup
    """

    # =========================
    # 1) LISTADO DE BACKUPS ACTIVOS (PAGINADO)
    # =========================
    backups_qs = (
        Backups.objects
        .filter(b_estado="activo")
        .select_related("b_fce", "b_disco")
        .order_by("-b_id")
    )

    page_backups = request.GET.get("page", 1)
    paginator_backups = Paginator(backups_qs, 10)
    backups = paginator_backups.get_page(page_backups)

    # =========================
    # 2) FUNCIONARIOS CON EQUIPO SIN BACKUP (CONTEO REAL + PAGINACIÓN)
    # =========================
    fce_con_backup_ids = (
        Backups.objects
        .filter(b_estado="activo")
        .values_list("b_fce_id", flat=True)
        .distinct()
    )

    fce_sin_backup_qs = (
        Funcionarioconequipo.objects
        .filter(fce_estado="activo")
        .exclude(fce_id__in=fce_con_backup_ids)
        .select_related("fce_fun")
        .order_by("fce_fun")  # si quieres por nombre real, cámbialo al campo exacto
    )

    total_sin_backup = fce_sin_backup_qs.count()

    page_sin = request.GET.get("page_sin", 1)
    paginator_sin = Paginator(fce_sin_backup_qs, 10)
    funcionarios_sin_backup = paginator_sin.get_page(page_sin)

    # =========================
    # 3) FORMULARIOS
    # =========================
    form2 = BackupsProcesoForm()

    # para el select de "Desactivar Backup": necesitamos TODOS (o al menos activos)
    # pero NO debe depender del paginator "backups" (solo trae 10)
    backups_activos_all = backups_qs.only("b_id", "b_nombre", "b_fce_id", "b_disco_id")

    # =========================
    # 4) POST: ASIGNAR / DESACTIVAR
    # =========================
    if request.method == "POST":
        if "add_bk" in request.POST:
            form2 = BackupsProcesoForm(request.POST)
            if form2.is_valid():
                obj = form2.save(commit=False)
                # fuerza estado activo si tu form no lo hace
                obj.b_estado = "activo"
                obj.save()
                messages.success(request, "✅ Backup asignado exitosamente")
                return redirect("bk-asignacion")
            else:
                messages.error(request, "❌ Error al asignar backup. Verifica los datos.")

        elif "delete_bk" in request.POST:
            bk_id = request.POST.get("bk_id")
            bk = get_object_or_404(Backups, pk=bk_id)
            try:
                bk.b_estado = "inactivo"
                bk.save(update_fields=["b_estado"])
                messages.success(request, "✅ Backup desactivado exitosamente")
            except Exception:
                messages.error(request, "❌ Error al desactivar el backup")
            return redirect("bk-asignacion")

    # =========================
    # 5) CONTEXT
    # =========================
    context = {
        "backups": backups,
        "paginator_backups": paginator_backups,

        "funcionarios_sin_backup": funcionarios_sin_backup,
        "paginator_sin": paginator_sin,
        "total_sin_backup": total_sin_backup,

        "form2": form2,
        "backups_activos_all": backups_activos_all,
    }

    return render(request, "backups/gestion_backups.html", context)






# DARLE/ELIMINAR UN BACKUP A UN FUNCIONARIO CON EQUIPO

# AÑADIR/ELIMINAR FUNCIONARIO CON EQUIPO


@login_required
@user_passes_test(is_swap_informatica)
def bk_configuracion(request):
    """Configuración de Equipos - Funcionarios con Equipo y Funcionarios con paginación"""
    
    from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
    
    # Obtener datos
    fce_list = Funcionarioconequipo.objects.filter(fce_estado='activo').order_by('fce_fun')
    funcionarios_list = Funcionarios.objects.filter(fun_estado='activo').order_by('fun_id')
    
    # Paginación para Funcionarios con Equipo
    page_fce = request.GET.get('page_fce', 1)
    paginator_fce = Paginator(fce_list, 10)  # 10 items por página
    try:
        fce = paginator_fce.page(page_fce)
    except PageNotAnInteger:
        fce = paginator_fce.page(1)
    except EmptyPage:
        fce = paginator_fce.page(paginator_fce.num_pages)
    
    # Paginación para Funcionarios
    page_func = request.GET.get('page_func', 1)
    paginator_func = Paginator(funcionarios_list, 10)  # 10 items por página
    try:
        funcionarios = paginator_func.page(page_func)
    except PageNotAnInteger:
        funcionarios = paginator_func.page(1)
    except EmptyPage:
        funcionarios = paginator_func.page(paginator_func.num_pages)
    
    # Forms
    form3 = fceForm()
    
    # Procesar formularios
    if request.method == "POST":
        if "add_fce" in request.POST:
            form3 = fceForm(request.POST, request.FILES)
            if form3.is_valid():
                form3.save()
                messages.success(request, '✅ Funcionario con Equipo agregado exitosamente')
                return redirect('config-equipos')
            else:
                messages.error(request, '❌ Error al agregar. Verifica los datos.')
        
        elif "delete_fce" in request.POST:
            fce_id = request.POST.get("fce_id")
            fce_obj = get_object_or_404(Funcionarioconequipo, pk=fce_id)
            try:
                fce_obj.fce_estado = "inactivo"
                fce_obj.save()
                messages.success(request, '✅ Registro desactivado exitosamente')
            except:
                messages.error(request, '❌ Error al desactivar el registro')
            return redirect('config-equipos')
    
    context = {
        'fce': fce,
        'funcionarios': funcionarios,
        'form3': form3,
        'paginator_fce': paginator_fce,
        'paginator_func': paginator_func,
    }
    return render(request, 'backups/configuracion_de_equipos.html', context)


# EXXCEEEELLL


@login_required
@user_passes_test(is_swap_informatica)
def exportar_excel_backups(request):
    # =========================
    # 1) FILTROS (IGNORA PAGINACIÓN)
    # =========================
    fecha_filtro  = (request.GET.get("fecha") or "").strip()
    fecha_desde   = (request.GET.get("fecha_desde") or "").strip()
    fecha_hasta   = (request.GET.get("fecha_hasta") or "").strip()
    backup_filtro = (request.GET.get("backup") or "").strip()

    backup_id = int(backup_filtro) if backup_filtro.isdigit() else None

    filtros = Q()

    if fecha_filtro:
        d = parse_date(fecha_filtro)
        if d:
            filtros &= Q(bh_fecha=d)
    else:
        if fecha_desde:
            d = parse_date(fecha_desde)
            if d:
                filtros &= Q(bh_fecha__gte=d)
        if fecha_hasta:
            d = parse_date(fecha_hasta)
            if d:
                filtros &= Q(bh_fecha__lte=d)

    if backup_id is not None:
        filtros &= Q(bh_bp__bp_b_id=backup_id)

    # ✅ trae TODOS los registros filtrados (sin paginación)
    qs = Backupshechos.objects.filter(filtros).order_by("-bh_fecha", "-bh_id")

    # =========================
    # 2) EXCEL
    # =========================
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Backups"

    # =========================
    # 3) ESTILOS
    # =========================
    def add_style(style: NamedStyle):
        existing = {s.name for s in wb.named_styles if hasattr(s, "name")}
        if style.name not in existing:
            wb.add_named_style(style)

    # Header
    header_style = NamedStyle(name="header_style_bk")
    header_style.font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
    header_style.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_style.border = Border(
        left=Side(style="medium"),
        right=Side(style="medium"),
        top=Side(style="medium"),
        bottom=Side(style="medium"),
    )
    header_style.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Data normal
    data_style = NamedStyle(name="data_style_bk")
    data_style.font = Font(name="Calibri", size=11, color="000000")
    data_style.border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    data_style.alignment = Alignment(vertical="center")

    # Data alternado
    alternate_style = NamedStyle(name="alternate_style_bk")
    alternate_style.fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    alternate_style.font = Font(name="Calibri", size=11, color="000000")
    alternate_style.border = data_style.border
    alternate_style.alignment = data_style.alignment

    # ✅ Estilos de FECHA (CLAVE: incluyen number_format y no se pisan)
    data_date_style = NamedStyle(name="data_date_style_bk")
    data_date_style.font = data_style.font
    data_date_style.border = data_style.border
    data_date_style.alignment = data_style.alignment
    data_date_style.number_format = "dd/mm/yyyy"

    alternate_date_style = NamedStyle(name="alternate_date_style_bk")
    alternate_date_style.fill = alternate_style.fill
    alternate_date_style.font = alternate_style.font
    alternate_date_style.border = alternate_style.border
    alternate_date_style.alignment = alternate_style.alignment
    alternate_date_style.number_format = "dd/mm/yyyy"

    for st in (header_style, data_style, alternate_style, data_date_style, alternate_date_style):
        add_style(st)

    # =========================
    # 4) ENCABEZADOS
    # =========================
    headers = ["ID", "Fecha (backup)", "Backup"]
    ws.append(headers)
    for col in range(1, 4):
        ws.cell(row=1, column=col).style = header_style

    ws.freeze_panes = "A2"

    # =========================
    # 5) DATOS
    # =========================
    def as_date(value):
        if not value:
            return None
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        if isinstance(value, str):
            return parse_date(value)
        return None

    # Evaluamos una vez (también evita qs.count() extra)
    rows = list(qs)

    for i, b in enumerate(rows, start=2):
        even = (i % 2 == 0)

        # ID
        c_id = ws.cell(row=i, column=1, value=b.bh_id)
        c_id.style = data_style if even else alternate_style

        # Fecha real del backup (bh_fecha)
        fecha_real = as_date(b.bh_fecha)
        c_date = ws.cell(row=i, column=2, value=fecha_real)
        c_date.style = data_date_style if even else alternate_date_style
        # (opcional) refuerzo: si hay fecha, marcar como tipo fecha
        if fecha_real is not None:
            c_date.data_type = "d"

        # Backup
        c_bk = ws.cell(row=i, column=3, value=str(b.bh_bp) if b.bh_bp else "(sin relación)")
        c_bk.style = data_style if even else alternate_style

    # =========================
    # 6) FORMATO / AUTO FILTER
    # =========================
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 60

    last_row = len(rows) + 1
    ws.auto_filter.ref = f"A1:C{max(last_row, 2)}"

    # ✅ Condicional solo si querés marcar FUTURAS
    future_date_rule = CellIsRule(
        operator="greaterThan",
        formula=["TODAY()"],
        stopIfTrue=True,
        font=Font(color="FF0000", italic=True),
    )
    ws.conditional_formatting.add(f"B2:B{max(last_row, 2)}", future_date_rule)

    # =========================
    # 7) RESPUESTA
    # =========================
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="backups_report.xlsx"'
    wb.save(response)
    return response








# lleva a otra pagina de backups hechos
@login_required
@user_passes_test(is_swap_informatica)
def backups_hechos(request):
    # ===== POST: eliminar =====
    if request.method == "POST":
        delete_ids = request.POST.getlist("delete_ids")
        if delete_ids:
            Backupshechos.objects.filter(bh_id__in=delete_ids).delete()

        # Mantener filtros/página al volver
        redirect_url = request.POST.get("redirect_to") or "bk-hechos"
        return redirect(redirect_url)

    # ===== GET: limpiar =====
    if "limpiar" in request.GET:
        return redirect("bk-hechos")

    # ===== GET: filtros =====
    fecha_filtro  = (request.GET.get("fecha") or "").strip()
    fecha_desde   = (request.GET.get("fecha_desde") or "").strip()
    fecha_hasta   = (request.GET.get("fecha_hasta") or "").strip()
    backup_filtro = (request.GET.get("backup") or "").strip()

    backup_id = int(backup_filtro) if backup_filtro.isdigit() else None

    filtros = Q()

    # Fecha exacta tiene prioridad
    if fecha_filtro:
        filtros &= Q(bh_fecha=parse_date(fecha_filtro))
    else:
        if fecha_desde:
            filtros &= Q(bh_fecha__gte=parse_date(fecha_desde))
        if fecha_hasta:
            filtros &= Q(bh_fecha__lte=parse_date(fecha_hasta))

    if backup_id is not None:
        filtros &= Q(bh_bp__bp_b_id=backup_id)

    # Query base
    qs = Backupshechos.objects.filter(filtros).order_by("-bh_fecha", "-bh_id")

    # Paginación (ajusta a tu gusto)
    per_page = 25
    paginator = Paginator(qs, per_page)
    page_number = request.GET.get("page") or 1
    page_obj = paginator.get_page(page_number)

    # Esto para el select del filtro
    backups = Backups.objects.all().order_by("b_id")

    return render(request, "backups/backups_hechos.html", {
        # 👇 ahora iteras page_obj, no qs
        "page_obj": page_obj,
        "bkdone": page_obj.object_list,  # por compatibilidad si ya lo usas
        "backups": backups,

        "fecha_filtro": fecha_filtro,
        "fecha_desde": fecha_desde,
        "fecha_hasta": fecha_hasta,
        "backup_filtro": str(backup_id) if backup_id is not None else "",

        "per_page": per_page,
    })



# ayuda a agregar los backupshechos
#@login_required
#@user_passes_test(is_swap_informatica)
# Los decoradores @login_required y @user_passes_test solo deben usarse 
# en vistas (funciones que reciben request). Tu función obtener_fecha_del_dia() e
# s una función utilitaria, no una vista, y no tiene ni necesita acceso a request.
def obtener_fecha_del_dia(dia_nombre, referencia):
    dias_semana = ["Lunes", "Martes", "Miercoles", "Jueves", "Viernes", "Sabado", "Domingo"]
    dia_actual = referencia.weekday()  # 0 = Lunes, 4 = Viernes
    dia_objetivo = dias_semana.index(dia_nombre)

    # Calcular la diferencia de días
    if dia_objetivo <= dia_actual:
        diferencia = dia_objetivo - dia_actual
    else:
        diferencia = dia_objetivo - dia_actual - 7

    return referencia + timedelta(days=diferencia)

# Se agregan los backups hechos (configuracion de horario para apretar el boton)
@user_passes_test(is_swap_informatica)
@login_required
def agregar_backups_hechos(request):
    fecha_actual = now()
    dia_actual = fecha_actual.strftime("%A")
    hora_actual = fecha_actual.strftime("%H:%M")

    # Verificar si está entre el horario permitido
    if (dia_actual == "Friday" and hora_actual >= "15:00") or dia_actual in ["Saturday", "Sunday"]:
        estado_finalizado = Backupsestado.objects.filter(be_estado="Finalizado✅").first()
        estado_pendiente = Backupsestado.objects.filter(be_estado="Pendiente⚠️").first()

        if estado_finalizado and estado_pendiente:
            procesos_finalizados = Backupsproceso.objects.filter(bp_be=estado_finalizado)

            if procesos_finalizados.exists():
                for proceso in procesos_finalizados:
                    backup_dia = proceso.bp_bd.bd_dia  # "Lunes", "Martes", etc.
                    fecha_backup = obtener_fecha_del_dia(backup_dia, fecha_actual.date())

                    # Guardar el backup hecho
                    Backupshechos.objects.create(
                        bh_fecha=fecha_backup,
                        bh_bp=proceso
                    )

                    # Cambiar el estado a "Pendiente⚠️"
                    proceso.bp_be = estado_pendiente
                    proceso.save()

    return redirect("backups")
# ==================== BACKUPS ====================



# ==================== INVENTARIO ====================
# INVENTARIO
@login_required
@user_passes_test(is_swap_informatica)
def inventario(request):
    inventario_informatica = Inventarioinformatica.objects.all()
    form = rakForm()
    form2 = StatusForm()
    form3 = ItemForm()
    form4 = NewArticleForm()


    # Obtener listas para los filtros
    categorias = InventarioinformaticaCategoria.objects.all()
    depositos = InventarioinformaticaDeposito.objects.all()
    estados = InventarioinformaticaEstado.objects.all()

    # Obtener parámetros del filtro
    articulo_filtro = request.GET.get('articulo', '').strip()
    deposito_filtro = request.GET.get('deposito', '').strip()
    estado_filtro = request.GET.get('estado', '').strip()

    # Aplicar filtros
    filtros = Q()
    if articulo_filtro:
        filtros &= Q(ii_iic_id=articulo_filtro)
    if deposito_filtro:
        filtros &= Q(ii_iid_id=deposito_filtro)
    if estado_filtro:
        filtros &= Q(ii_iie_id=estado_filtro)

    inventario_informatica = Inventarioinformatica.objects.filter(filtros)

    # datos para las ultimas 3 tablas
    items = InventarioinformaticaCategoria.objects.all()
    raks = InventarioinformaticaDeposito.objects.all()
    status = InventarioinformaticaEstado.objects.all()

    return render(request, "inventario.html", {
        'inventario_informatica': inventario_informatica,
        'form':form,
        'form2':form2,
        'form3': form3,
        'form4': form4,

        'categorias': categorias,
        'depositos': depositos,
        'estados': estados,
        'articulo_filtro': articulo_filtro,
        'deposito_filtro': deposito_filtro,
        'estado_filtro': estado_filtro,

        'items': items,
        'raks': raks,
        'status': status,
    })

# se añaden raks o depositos para inventario
@login_required
@user_passes_test(is_swap_informatica)
def agregar_rak(request):
    if request.method == "POST":
        form = rakForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect("inventario")
    else:
        form = rakForm()  # Instancia vacía si es GET
    return redirect("inventario")

# se añaden estados para inventario
@login_required
@user_passes_test(is_swap_informatica)
def agregar_status(request):
    if request.method == "POST":
        form2 = StatusForm(request.POST)
        if form2.is_valid():
            form2.save()
            return redirect("inventario")
    else:
        form2 = StatusForm()  # Instancia vacía si es GET
    return redirect("inventario")

# se añade un item
@login_required
@user_passes_test(is_swap_informatica)
def agregar_item(request):
    if request.method == "POST":
        form3 = ItemForm(request.POST)
        if form3.is_valid():
            form3.save()
            return redirect("inventario")
    else:
        form3 = ItemForm()  # Instancia vacía si es GET
    return redirect("inventario")

# se añade un nuevo articulo
@login_required
@user_passes_test(is_swap_informatica)
def agregar_articulo(request):
    if request.method == "POST":
        form4 = NewArticleForm(request.POST)
        if form4.is_valid():
            article = form4.save(commit=False)  # No guardamos aún en la BD
            article.ii_fecha = now().date()  # Asignamos la fecha actual
            article.save()  # Guardamos en la BD
            return redirect("inventario")
    else:
        form4 = NewArticleForm()

    return redirect("inventario")

# se edita un articulo
@csrf_exempt
@login_required
@user_passes_test(is_swap_informatica)
def guardar_cantidad(request):
    if request.method == "POST":
        data = json.loads(request.body)
        try:
            item = Inventarioinformatica.objects.get(ii_id=data["id"])
            item.ii_cantidad = data["cantidad"]
            item.ii_fecha = timezone.now().date()  # 👈 actualizar la fecha
            item.save()
            return JsonResponse({"status": "success"})
        except Inventarioinformatica.DoesNotExist:
            return JsonResponse({"status": "error", "message": "Artículo no encontrado"})
        
# Equipos
@login_required
@user_passes_test(is_swap_informatica)
def equipos(request):
    # Tabla General de Vistas
    equipos = Equipos.objects.select_related('eq_tipe').filter(
    Q(eq_tipe__tipe_nombre='PC') | Q(eq_tipe__tipe_nombre='Notebook'))
    
    upss = Equipos.objects.select_related('eq_tipe').filter(
    eq_tipe__tipe_nombre='UPS')
    
    impresoras = Equipos.objects.select_related('eq_tipe').filter(
        eq_tipe__tipe_nombre='Impresora')
    
    otros = Equipos.objects.select_related('eq_tipe').filter(
        Q(eq_tipe__tipe_nombre='Router') | Q(eq_tipe__tipe_nombre='DVR') | 
        Q(eq_tipe__tipe_nombre='Laser') | Q(eq_tipe__tipe_nombre='Scaner')
    )
    
    cosas_servidores = Equipos.objects.select_related('eq_tipe').filter(
        Q(eq_tipe__tipe_nombre__icontains='switch') |
        Q(eq_tipe__tipe_nombre__icontains='rack') |
        Q(eq_tipe__tipe_nombre__icontains='servidor')
    )
    
    monitores = Equipos.objects.select_related('eq_tipe').filter(
    eq_tipe__tipe_nombre='Monitor')
    
    # Tabla de Vistas de Mantenimientos
    # ===============================================================================================
    mantenimientos_equipos = Mantenimiento.objects.filter(
    Q(m_eq__eq_tipe__tipe_nombre="PC") | Q(m_eq__eq_tipe__tipe_nombre="Notebook")
    )
    
    # Filtros de mantenimientos
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    fecha_exacta = request.GET.get('fecha_exacta')
    tipo_mantenimiento = request.GET.get('tipo_mantenimiento')
    responsable = request.GET.get('responsable')
    equipo_id = request.GET.get('equipo_id')

    if fecha_exacta:
        mantenimientos_equipos = mantenimientos_equipos.filter(m_fecha=fecha_exacta)
    else:
        if fecha_desde:
            mantenimientos_equipos = mantenimientos_equipos.filter(m_fecha__gte=fecha_desde)
        if fecha_hasta:
            mantenimientos_equipos = mantenimientos_equipos.filter(m_fecha__lte=fecha_hasta)

    if tipo_mantenimiento:
        mantenimientos_equipos = mantenimientos_equipos.filter(m_mt__mt_nombre__iexact=tipo_mantenimiento)

    if responsable:
        mantenimientos_equipos = mantenimientos_equipos.filter(m_responsable__icontains=responsable)

    if equipo_id:
        mantenimientos_equipos = mantenimientos_equipos.filter(m_eq__eq_id=equipo_id)

    # ===============================================================================================
    
    mantenimientos_ups = Mantenimiento.objects.filter(
        m_eq__eq_tipe__tipe_nombre="UPS"
    )
    
    # Filtros para mantenimientos de UPS
    fecha_desde_ups = request.GET.get('fecha_desde_ups')
    fecha_hasta_ups = request.GET.get('fecha_hasta_ups')
    fecha_exacta_ups = request.GET.get('fecha_exacta_ups')
    tipo_mantenimiento_ups = request.GET.get('tipo_mantenimiento_ups')
    responsable_ups = request.GET.get('responsable_ups')
    equipo_id_ups = request.GET.get('equipo_id_ups')

    if fecha_exacta_ups:
        mantenimientos_ups = mantenimientos_ups.filter(m_fecha=fecha_exacta_ups)
    else:
        if fecha_desde_ups:
            mantenimientos_ups = mantenimientos_ups.filter(m_fecha__gte=fecha_desde_ups)
        if fecha_hasta_ups:
            mantenimientos_ups = mantenimientos_ups.filter(m_fecha__lte=fecha_hasta_ups)

    if tipo_mantenimiento_ups:
        mantenimientos_ups = mantenimientos_ups.filter(m_mt__mt_nombre__iexact=tipo_mantenimiento_ups)

    if responsable_ups:
        mantenimientos_ups = mantenimientos_ups.filter(m_responsable__icontains=responsable_ups)

    if equipo_id_ups:
        mantenimientos_ups = mantenimientos_ups.filter(m_eq__eq_id=equipo_id_ups)

    
    # ===============================================================================================
    
    mantenimientos_impresoras = Mantenimiento.objects.filter(
        m_eq__eq_tipe__tipe_nombre="Impresora"
    )
    
    # Filtros para mantenimientos de impresoras
    fecha_desde_imp = request.GET.get('fecha_desde_imp')
    fecha_hasta_imp = request.GET.get('fecha_hasta_imp')
    fecha_exacta_imp = request.GET.get('fecha_exacta_imp')
    tipo_mantenimiento_imp = request.GET.get('tipo_mantenimiento_imp')
    responsable_imp = request.GET.get('responsable_imp')
    equipo_id_imp = request.GET.get('equipo_id_imp')

    if fecha_exacta_imp:
        mantenimientos_impresoras = mantenimientos_impresoras.filter(m_fecha=fecha_exacta_imp)
    else:
        if fecha_desde_imp:
            mantenimientos_impresoras = mantenimientos_impresoras.filter(m_fecha__gte=fecha_desde_imp)
        if fecha_hasta_imp:
            mantenimientos_impresoras = mantenimientos_impresoras.filter(m_fecha__lte=fecha_hasta_imp)

    if tipo_mantenimiento_imp:
        mantenimientos_impresoras = mantenimientos_impresoras.filter(m_mt__mt_nombre__iexact=tipo_mantenimiento_imp)

    if responsable_imp:
        mantenimientos_impresoras = mantenimientos_impresoras.filter(m_responsable__icontains=responsable_imp)

    if equipo_id_imp:
        mantenimientos_impresoras = mantenimientos_impresoras.filter(m_eq__eq_id=equipo_id_imp)

    
    # ===============================================================================================
    
    mantenimientos_otros = Mantenimiento.objects.filter(
    Q(m_eq__eq_tipe__tipe_nombre__icontains="router") | Q(m_eq__eq_tipe__tipe_nombre__icontains="DVR") |
    Q(m_eq__eq_tipe__tipe_nombre__icontains="laser") | Q(m_eq__eq_tipe__tipe_nombre__icontains="Scaner")
    )
    
    # Filtros para mantenimientos de otros dispositivos
    fecha_desde_otros = request.GET.get('fecha_desde_otros')
    fecha_hasta_otros = request.GET.get('fecha_hasta_otros')
    fecha_exacta_otros = request.GET.get('fecha_exacta_otros')
    tipo_mantenimiento_otros = request.GET.get('tipo_mantenimiento_otros')
    responsable_otros = request.GET.get('responsable_otros')
    equipo_id_otros = request.GET.get('equipo_id_otros')

    if fecha_exacta_otros:
        mantenimientos_otros = mantenimientos_otros.filter(m_fecha=fecha_exacta_otros)
    else:
        if fecha_desde_otros:
            mantenimientos_otros = mantenimientos_otros.filter(m_fecha__gte=fecha_desde_otros)
        if fecha_hasta_otros:
            mantenimientos_otros = mantenimientos_otros.filter(m_fecha__lte=fecha_hasta_otros)

    if tipo_mantenimiento_otros:
        mantenimientos_otros = mantenimientos_otros.filter(m_mt__mt_nombre__iexact=tipo_mantenimiento_otros)

    if responsable_otros:
        mantenimientos_otros = mantenimientos_otros.filter(m_responsable__icontains=responsable_otros)

    if equipo_id_otros:
        mantenimientos_otros = mantenimientos_otros.filter(m_eq__eq_id=equipo_id_otros)

    
    # ===============================================================================================
    
    mantenimientos_servidores = Mantenimiento.objects.filter(
        Q(m_eq__eq_tipe__tipe_nombre__icontains="switch") | Q(m_eq__eq_tipe__tipe_nombre__icontains="rack") |
        Q(m_eq__eq_tipe__tipe_nombre__icontains="servidor")
    )

    # Filtros para mantenimientos de servidores
    fecha_desde_srv = request.GET.get('fecha_desde_srv')
    fecha_hasta_srv = request.GET.get('fecha_hasta_srv')
    fecha_exacta_srv = request.GET.get('fecha_exacta_srv')
    tipo_mantenimiento_srv = request.GET.get('tipo_mantenimiento_srv')
    responsable_srv = request.GET.get('responsable_srv')
    equipo_id_srv = request.GET.get('equipo_id_srv')

    if fecha_exacta_srv:
        mantenimientos_servidores = mantenimientos_servidores.filter(m_fecha=fecha_exacta_srv)
    else:
        if fecha_desde_srv:
            mantenimientos_servidores = mantenimientos_servidores.filter(m_fecha__gte=fecha_desde_srv)
        if fecha_hasta_srv:
            mantenimientos_servidores = mantenimientos_servidores.filter(m_fecha__lte=fecha_hasta_srv)

    if tipo_mantenimiento_srv:
        mantenimientos_servidores = mantenimientos_servidores.filter(m_mt__mt_nombre__iexact=tipo_mantenimiento_srv)

    if responsable_srv:
        mantenimientos_servidores = mantenimientos_servidores.filter(m_responsable__icontains=responsable_srv)

    if equipo_id_srv:
        mantenimientos_servidores = mantenimientos_servidores.filter(m_eq__eq_id=equipo_id_srv)

    

    # Filtros para cosas de servidores
    tipo_equipo_servidor = request.GET.get('tipo_equipo_servidor')
    departamento_servidor = request.GET.get('departamento_servidor')
    usuario_servidor = request.GET.get('usuario_servidor')

    cosas_servidores = Equipos.objects.select_related('eq_tipe')

    # Filtrar por tipo
    if tipo_equipo_servidor:
        cosas_servidores = cosas_servidores.filter(eq_tipe__tipe_nombre__iexact=tipo_equipo_servidor)
    else:
        cosas_servidores = cosas_servidores.filter(
            Q(eq_tipe__tipe_nombre__icontains='switch') |
            Q(eq_tipe__tipe_nombre__icontains='rack') |
            Q(eq_tipe__tipe_nombre__icontains='servidor')
        )

    # Filtrar por departamento
    if departamento_servidor:
        cosas_servidores = cosas_servidores.filter(eq_dpto=departamento_servidor)

    # Filtrar por usuario
    if usuario_servidor:
        cosas_servidores = cosas_servidores.filter(eq_usuario__icontains=usuario_servidor)


    # ===============================================================================================

    mantenimientos_monitores = Mantenimiento.objects.filter(
        m_eq__eq_tipe__tipe_nombre__icontains="monitor"
    )
    
    # Filtros para mantenimientos de monitores
    fecha_desde_monitor = request.GET.get('fecha_desde_monitor')
    fecha_hasta_monitor = request.GET.get('fecha_hasta_monitor')
    fecha_exacta_monitor = request.GET.get('fecha_exacta_monitor')
    tipo_mantenimiento_monitor = request.GET.get('tipo_mantenimiento_monitor')
    responsable_monitor = request.GET.get('responsable_monitor')
    equipo_id_monitor = request.GET.get('equipo_id_monitor')

    if fecha_exacta_monitor:
        mantenimientos_monitores = mantenimientos_monitores.filter(m_fecha=fecha_exacta_monitor)
    else:
        if fecha_desde_monitor:
            mantenimientos_monitores = mantenimientos_monitores.filter(m_fecha__gte=fecha_desde_monitor)
        if fecha_hasta_monitor:
            mantenimientos_monitores = mantenimientos_monitores.filter(m_fecha__lte=fecha_hasta_monitor)

    if tipo_mantenimiento_monitor:
        mantenimientos_monitores = mantenimientos_monitores.filter(m_mt__mt_nombre__iexact=tipo_mantenimiento_monitor)

    if responsable_monitor:
        mantenimientos_monitores = mantenimientos_monitores.filter(m_responsable__icontains=responsable_monitor)

    if equipo_id_monitor:
        mantenimientos_monitores = mantenimientos_monitores.filter(m_eq__eq_id=equipo_id_monitor)

    
    # ===============================================================================================
    
    # filtro equipos
    tipo_equipo = request.GET.get('tipo_equipo')
    departamento = request.GET.get('departamento')
    usuario = request.GET.get('usuario')
    departamentos = Departamentos.objects.all()
    if tipo_equipo:
        equipos = equipos.filter(eq_tipe__tipe_nombre__iexact=tipo_equipo)
    
    if departamento:
        equipos = equipos.filter(eq_dpto__dpto_id=departamento)

    if usuario:
        equipos = equipos.filter(responsable__icontains=usuario)
    
    # filtro ups's
    if departamento:
        upss = upss.filter(eq_dpto__dpto_id=departamento)
        
    if usuario:
        upss = upss.filter(responsable__icontains=usuario)#

    # filtro impresoras
    if departamento:
        impresoras = impresoras.filter(eq_dpto__dpto_id=departamento)

    if usuario:
        impresoras = impresoras.filter(responsable__icontains=usuario)
        
    # filtro otros
    tipo_equipo_otros = request.GET.get('tipo_equipo_otros')
    departamento_otros = request.GET.get('departamento_otros')
    usuario_otros = request.GET.get('usuario_otros')

    if tipo_equipo_otros:
        otros = otros.filter(eq_tipe__tipe_nombre__iexact=tipo_equipo_otros)

    if departamento_otros:
        otros = otros.filter(eq_dpto__dpto_id=departamento_otros)

    if usuario_otros:
        otros = otros.filter(responsable=usuario_otros)
        
    # filtro cosas_servidores
    tipo_equipo = request.GET.get('tipo_equipo_servidor')

    if tipo_equipo:
        if tipo_equipo == 'servidor':
            cosas_servidores = cosas_servidores.filter(eq_tipe__tipe_nombre__icontains='servidor')
        else:
            cosas_servidores = cosas_servidores.filter(eq_tipe__tipe_nombre__iexact=tipo_equipo)

    # filtro monitores
    departamento_monitor = request.GET.get('departamento_monitor')
    usuario_monitor = request.GET.get('usuario_monitor')

    if departamento_monitor:
        monitores = monitores.filter(eq_dpto__dpto_id=departamento_monitor)

    if usuario_monitor:
        monitores = monitores.filter(responsable=usuario_monitor)

        
    return render(request, "equipos.html",{
        'equipos': equipos,
        'upss': upss,
        'impresoras': impresoras,
        'otros': otros,
        'cosas_servidores': cosas_servidores,
        'monitores': monitores,
        
        'mantenimientos_equipos':mantenimientos_equipos,
        'mantenimientos_ups':mantenimientos_ups,
        'mantenimientos_impresoras':mantenimientos_impresoras,
        'mantenimientos_otros':mantenimientos_otros,
        'mantenimientos_servidores':mantenimientos_servidores,
        'mantenimientos_monitores':mantenimientos_monitores,
        
        # NUEVO: enviar departamentos para el filtro
        'departamentos': departamentos,
    })
    
# AÑADIR EQUIPOS
@login_required
@user_passes_test(is_swap_informatica)
def NuevoEquipo(request):
    form_type = request.GET.get('form', 'equipo')  # por defecto equipo

    # Diccionario para mapear los tipos a los formularios
    form_classes = {
        'equipo': NewEquipo,
        'ups': NewUps,
        'impresora': NewImpresora,
        'otros': NewOtros,
        'servidor': NewCosasServidor,
        'monitor': NewMonitor,
    }

    FormClass = form_classes.get(form_type)

    if FormClass is None:
        return render(request, '404.html')  # o redirigir a una página de error personalizada

    if request.method == 'POST':
        form = FormClass(request.POST)
        if form.is_valid():
            form.save()
            return redirect('equipos')  # cambialo por la URL que tengas
    else:
        form = FormClass()

    return render(request, 'add.html', {'form': form, 'form_type': form_type})
# ==================== INVENTARIO ====================



# ========== MANTENIMIENTO ====================
# MANTENIMIENTO
from django.db.models.functions import ExtractMonth

@login_required
@user_passes_test(is_swap_informatica)
def mantenimiento(request):
    mantenimientos_enero = Mantenimientocalendario.objects.annotate(
        mes=ExtractMonth('mc_fecha')).filter(mes=1).order_by('mc_fecha')
    
    mantenimientos_febrero = Mantenimientocalendario.objects.annotate(
        mes=ExtractMonth('mc_fecha')).filter(mes=2).order_by('mc_fecha')
    
    mantenimientos_marzo = Mantenimientocalendario.objects.annotate(
        mes=ExtractMonth('mc_fecha')).filter(mes=3).order_by('mc_fecha')
    
    mantenimientos_abril = Mantenimientocalendario.objects.annotate(
        mes=ExtractMonth('mc_fecha')).filter(mes=4).order_by('mc_fecha')
    
    mantenimientos_mayo = Mantenimientocalendario.objects.annotate(
        mes=ExtractMonth('mc_fecha')).filter(mes=5).order_by('mc_fecha')
    
    mantenimientos_junio = Mantenimientocalendario.objects.annotate(
        mes=ExtractMonth('mc_fecha')).filter(mes=6).order_by('mc_fecha')
    
    mantenimientos_julio = Mantenimientocalendario.objects.annotate(
        mes=ExtractMonth('mc_fecha')).filter(mes=7).order_by('mc_fecha')
    
    mantenimientos_agosto = Mantenimientocalendario.objects.annotate(
        mes=ExtractMonth('mc_fecha')).filter(mes=8).order_by('mc_fecha')
    
    mantenimientos_septiembre = Mantenimientocalendario.objects.annotate(
        mes=ExtractMonth('mc_fecha')).filter(mes=9).order_by('mc_fecha')
    
    mantenimientos_octubre = Mantenimientocalendario.objects.annotate(
        mes=ExtractMonth('mc_fecha')).filter(mes=10).order_by('mc_fecha')
    
    mantenimientos_noviembre = Mantenimientocalendario.objects.annotate(
        mes=ExtractMonth('mc_fecha')).filter(mes=11).order_by('mc_fecha')
    
    mantenimientos_diciembre = Mantenimientocalendario.objects.annotate(
        mes=ExtractMonth('mc_fecha')).filter(mes=12).order_by('mc_fecha')
    
    return render(request, "mantenimiento.html", {
        'mantenimientos_enero': mantenimientos_enero,
        'mantenimientos_febrero': mantenimientos_febrero,
        'mantenimientos_marzo': mantenimientos_marzo,
        'mantenimientos_abril': mantenimientos_abril,
        'mantenimientos_mayo': mantenimientos_mayo,
        'mantenimientos_junio': mantenimientos_junio,
        'mantenimientos_julio': mantenimientos_julio,
        'mantenimientos_agosto': mantenimientos_agosto,
        'mantenimientos_septiembre': mantenimientos_septiembre,
        'mantenimientos_octubre': mantenimientos_octubre,
        'mantenimientos_noviembre': mantenimientos_noviembre,
        'mantenimientos_diciembre': mantenimientos_diciembre,
    })


@require_POST
@login_required
@user_passes_test(is_swap_informatica)
def cambiar_estado_mantenimiento(request, mc_id):
    nuevo_estado = request.POST.get('estado')
    responsable = request.POST.get('responsable')  # viene vacío si no es finalizado

    mantenimiento = get_object_or_404(Mantenimientocalendario, mc_id=mc_id)

    estado_obj, _ = Mantenimientoestado.objects.get_or_create(me_estado=nuevo_estado)
    mantenimiento.mc_me = estado_obj
    mantenimiento.save()

    if nuevo_estado == "Finalizado✅" and responsable:
        crear_mantenimiento_desde_calendario(mantenimiento, responsable)

    return redirect('mantenimiento')

@login_required
@user_passes_test(is_swap_informatica)
def addmantenimiento(request):
    if request.method == 'POST':
        form = MantenimientoForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('addmantenimiento')  # Redirige a donde necesites
    else:
        form = MantenimientoForm()
    return render(request, 'addmantenimiento.html',{
        'form':form,
    })

# ========== MANTENIMIENTO ====================

# bk = Backups.objects.filter(b_estado='activo')
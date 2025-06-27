from django.shortcuts import render, redirect, get_object_or_404,HttpResponse
from .forms import *
from swap_home.models import *
from swap_informatica.models import *
from swap_porteria.models import *
from swap_serviciotecnico.models import *
from swap_tecnico.models import *
from swap_tthh.models import *
from django.db.models import Q
import openpyxl
from openpyxl.utils import get_column_letter
from django.contrib.auth.decorators import login_required, user_passes_test
from datetime import datetime
from openpyxl.styles import (
    Font, PatternFill, Border, Side, Alignment,
    NamedStyle, numbers, GradientFill
)
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from openpyxl import Workbook
# Create your views here.

def is_swap_tthh(user):
    return user.groups.filter(name='TTHH').exists()

# Vista Principal
@login_required
@user_passes_test(is_swap_tthh)
def tthh(request):
    tarea = Tareas.objects.filter(Q(tarea_dpto__dpto_nombre="TTHH") | Q(tarea_dpto__dpto_nombre="General"))
    tarea_dia = Tareadia.objects.all()
    tarea_estado = Tareaestado.objects.all()
    agregar_form = TareaForm()
    eliminar_form = EliminarTareaForm()
    
    if request.method == "POST":
        if "agregar_tarea" in request.POST:
            agregar_form = TareaForm(request.POST)  
            if agregar_form.is_valid():
                tarea = agregar_form.save(commit=False)
                #if tarea.tarea_td == "General":
                if tarea.tarea_td.td_dia == "General":
                    departamento_informatica = Departamentos.objects.get(dpto_nombre="General")
                else:
                    departamento_informatica = Departamentos.objects.get(dpto_nombre="TTHH")
                tarea.tarea_dpto = departamento_informatica
                tarea.save()
                return redirect("tthh")
        elif "eliminar_tarea" in request.POST:
            eliminar_form = EliminarTareaForm(request.POST)
            if eliminar_form.is_valid():
                tarea = eliminar_form.cleaned_data["tarea_titulo"]
                tarea.delete()
                return redirect("tthh")
    
    return render(request, 'tthh.html',{
        'tarea': tarea,
        'dia': tarea_dia,
        'estado': tarea_estado,
        'agregar_form': agregar_form,
        'eliminar_form': eliminar_form,
    })
    
# Actualizacion Estado de Tarea
@login_required
@user_passes_test(is_swap_tthh)
def actualizar_estado(request, tarea_id):
    if request.method == "POST":
        tarea = get_object_or_404(Tareas, pk=tarea_id)
        # Obtener el nuevo estado desde el formulario
        nuevo_estado = request.POST.get("estado")
        # Buscar el estado en la base de datos
        estado_obj = get_object_or_404(Tareaestado, te_estado=nuevo_estado)
        # Actualizar la tarea con el nuevo estado
        tarea.tarea_te = estado_obj
        tarea.save()
    return redirect("tthh")  # Redirige a la misma página después de actualizar

# Vista de Todas las Entradas Venidas desde Porteria
@login_required
@user_passes_test(is_swap_tthh)
def registro_entrada(request):
    #order_by('e_fecha') ordenar de menor fecha a mayor fecha(ascendente)
    funcionarios = Entrada.objects.filter(e_ee__ee_estado="Finalizado✅", e_fun__isnull=False).order_by('-e_fecha')

    # Obtener filtros del formulario
    fecha_exacta = request.GET.get('fecha_exacta')
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    funcionario_id = request.GET.get('funcionario')
    solo_tarde = request.GET.get('solo_tarde')  # checkbox o valor booleano

    # Aplicar filtros
    if fecha_exacta:
        funcionarios = funcionarios.filter(e_fecha=fecha_exacta)
    if fecha_desde:
        funcionarios = funcionarios.filter(e_fecha__gte=fecha_desde)
    if fecha_hasta:
        funcionarios = funcionarios.filter(e_fecha__lte=fecha_hasta)
    if funcionario_id:
        funcionarios = funcionarios.filter(e_fun=funcionario_id)
    if solo_tarde == "1":
        funcionarios = funcionarios.filter(e_entrada__gt="07:35:00")

    lista_funcionarios = Funcionarios.objects.all()



    return render(request, "registro_entrada.html", {
        'funcionarios': funcionarios,
        'lista_funcionarios': lista_funcionarios,
    })

# Excel de Funcionarios
@login_required
@user_passes_test(is_swap_tthh)
def exportar_excel_funcionario(request):
    # ========== FILTRADO DE DATOS ==========
    funcionarios = Entrada.objects.filter(e_ee__ee_estado="Finalizado✅", e_fun__isnull=False).order_by('-e_fecha')

    # Aplicar filtros (MANTENIENDO TU LÓGICA ORIGINAL)
    fecha_exacta = request.GET.get('fecha_exacta')
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    funcionario_id = request.GET.get('funcionario')
    solo_tarde = request.GET.get('solo_tarde')

    if fecha_exacta:
        funcionarios = funcionarios.filter(e_fecha=fecha_exacta)
    if fecha_desde:
        funcionarios = funcionarios.filter(e_fecha__gte=fecha_desde)
    if fecha_hasta:
        funcionarios = funcionarios.filter(e_fecha__lte=fecha_hasta)
    if funcionario_id:
        funcionarios = funcionarios.filter(e_fun=funcionario_id)
    if solo_tarde == "1":
        funcionarios = funcionarios.filter(e_entrada__gt="07:35:00")

    # ========== CONFIGURACIÓN EXCEL ==========
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Registro Entradas"

    # --- Configuración de estilos ---
    COLOR_HEADER = "4B0082"  # Índigo
    COLOR_TARDE = "FF6B6B"   # Rojo suave
    COLOR_FONDO = "F9F9F9"   # Gris claro

    # Estilo encabezados
    header_style = NamedStyle(name="header_style")
    header_style.font = Font(bold=True, color="FFFFFF", size=12)
    header_style.fill = PatternFill(start_color=COLOR_HEADER, fill_type="solid")
    header_style.alignment = Alignment(horizontal="center", vertical="center")
    header_style.border = Border(bottom=Side(style="thick"))

    # Estilo celdas normales
    cell_style = NamedStyle(name="cell_style")
    cell_style.font = Font(size=11)
    cell_style.alignment = Alignment(vertical="center")
    cell_style.border = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD")
    )

    # Estilo llegadas tarde
    late_style = NamedStyle(name="late_style")
    late_style.font = Font(color=COLOR_TARDE, bold=True)
    late_style.fill = PatternFill(start_color="FFF0F0", fill_type="solid")

    # Registrar estilos
    for style in [header_style, cell_style, late_style]:
        wb.add_named_style(style)

    # ========== ENCABEZADO ==========
    # Título principal
    ws.merge_cells('A1:F1')
    ws['A1'] = "REPORTE DE ASISTENCIA - TALENTO HUMANO"
    ws['A1'].font = Font(bold=True, size=14, color=COLOR_HEADER)
    ws['A1'].alignment = Alignment(horizontal="center")

    # Subtítulo
    ws.merge_cells('A2:F2')
    ws['A2'] = f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws['A2'].font = Font(italic=True, size=10, color="666666")
    ws['A2'].alignment = Alignment(horizontal="center")

    # Encabezados de columnas
    headers = ['FECHA', 'FUNCIONARIO', 'HORA ENTRADA', 'HORA SALIDA', 'ESTADO', 'COMENTARIOS']
    ws.append(headers)
    
    # Aplicar estilo a encabezados
    for col in range(1, len(headers)+1):
        cell = ws.cell(row=3, column=col)
        cell.style = header_style
        ws.column_dimensions[get_column_letter(col)].width = 20

    # ========== DATOS ==========
    # Inicializar row_idx (SOLUCIÓN AL ERROR)
    row_idx = 3  # Fila inicial (después de encabezados)

    if funcionarios.exists():  # Solo procesar si hay registros
        for row_idx, funcionario in enumerate(funcionarios, start=4):
            es_tarde = funcionario.e_entrada and funcionario.e_entrada.strftime('%H:%M:%S') > "07:35:00"
            
            ws.append([
                funcionario.e_fecha.strftime('%d/%m/%Y') if funcionario.e_fecha else '-',
                str(funcionario.e_fun) if funcionario.e_fun else '-',
                funcionario.e_entrada.strftime('%H:%M') if funcionario.e_entrada else '-',
                funcionario.e_salida.strftime('%H:%M') if funcionario.e_salida else '-',
                str(funcionario.e_ee) if funcionario.e_ee else '-',
                '-' if funcionario.e_comentario in [None, '', 'None'] else funcionario.e_comentario,
            ])

            # Aplicar estilos
            for col in range(1, len(headers)+1):
                cell = ws.cell(row=row_idx, column=col)
                cell.style = cell_style
                if es_tarde and col in [3, 4]:
                    cell.style = late_style

            # Efecto zebra
            if row_idx % 2 == 0:
                for col in range(1, len(headers)+1):
                    ws.cell(row=row_idx, column=col).fill = PatternFill(start_color=COLOR_FONDO, fill_type="solid")
    else:
        # Mensaje si no hay datos (OPCIONAL)
        ws.merge_cells('A4:F4')
        ws['A4'] = "No hay registros con los filtros actuales"
        ws['A4'].font = Font(color="FF0000", italic=True)
        ws['A4'].alignment = Alignment(horizontal="center")
        row_idx = 4  # Actualizar para el pie de página

    # ========== PIE DE PÁGINA ==========
    ws.merge_cells(f'A{row_idx+1}:F{row_idx+1}')
    ws[f'A{row_idx+1}'] = f"Total registros: {len(funcionarios)} | © {datetime.now().year} Departamento de Talento Humano"
    ws[f'A{row_idx+1}'].font = Font(italic=True, size=10, color="666666")
    ws[f'A{row_idx+1}'].alignment = Alignment(horizontal="center")
    ws[f'A{row_idx+1}'].fill = PatternFill(start_color="F0F0F0", fill_type="solid")

    # Congelar encabezados
    ws.freeze_panes = "A4"

    # ========== GENERAR RESPUESTA ==========
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename="reporte_asistencia.xlsx"'}
    )
    wb.save(response)
    return response

# Proveedores Ingresados
@login_required
@user_passes_test(is_swap_tthh)
def proveedores(request):
    # Filtro base: entradas finalizadas con proveedor no nulo
    proveedores_qs = Entrada.objects.filter(
        e_ee__ee_estado="Finalizado✅",
        e_prov__isnull=False
    ).select_related('e_prov', 'e_ee')

    # Obtener parámetros del GET
    proveedor_id = request.GET.get("proveedor")
    fecha_exacta = request.GET.get("fecha_exacta")
    fecha_inicio = request.GET.get("fecha_inicio")
    fecha_fin = request.GET.get("fecha_fin")

    # Filtro por proveedor (CORRECCIÓN IMPORTANTE)
    if proveedor_id and proveedor_id != "":
        proveedores_qs = proveedores_qs.filter(e_prov__prov_id=proveedor_id)

    # Filtros de fecha
    if fecha_exacta:
        proveedores_qs = proveedores_qs.filter(e_fecha=fecha_exacta)
    if fecha_inicio and fecha_fin:
        proveedores_qs = proveedores_qs.filter(e_fecha__range=[fecha_inicio, fecha_fin])
    elif fecha_inicio:
        proveedores_qs = proveedores_qs.filter(e_fecha__gte=fecha_inicio)
    elif fecha_fin:
        proveedores_qs = proveedores_qs.filter(e_fecha__lte=fecha_fin)

    # Obtener lista de proveedores activos para el dropdown
    lista_proveedores = Proveedor.objects.filter(
        prov_estado="activo"
    ).order_by('prov_nombre')

    return render(request, "proveedores.html", {
        'proveedores': proveedores_qs.order_by('-e_fecha'),
        'lista_proveedores': lista_proveedores,
    })
    
# Excel de Proveedores
@login_required
@user_passes_test(is_swap_tthh)
def exportar_excel_proveedores(request):
    queryset = Entrada.objects.filter(e_ee__ee_estado="Finalizado✅", e_prov__isnull=False)

    proveedor_id = request.GET.get("proveedor")
    fecha_exacta = request.GET.get("fecha_exacta")
    fecha_inicio = request.GET.get("fecha_inicio")
    fecha_fin = request.GET.get("fecha_fin")

    if proveedor_id:
        queryset = queryset.filter(e_prov_id=proveedor_id)
    if fecha_exacta:
        queryset = queryset.filter(e_fecha=fecha_exacta)
    if fecha_inicio and fecha_fin:
        queryset = queryset.filter(e_fecha__range=[fecha_inicio, fecha_fin])
    elif fecha_inicio:
        queryset = queryset.filter(e_fecha__gte=fecha_inicio)
    elif fecha_fin:
        queryset = queryset.filter(e_fecha__lte=fecha_fin)

    queryset = queryset.order_by('-e_fecha')

    # ========== CONFIGURACIÓN EXCEL ==========
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Registro Proveedores"

    # --- Configuración de estilos ---
    COLOR_HEADER = "2E5C8A"  # Azul profesional
    COLOR_FONDO = "F9F9F9"   # Gris claro
    COLOR_DESTACADO = "FFD700"  # Amarillo dorado para destacar

    # Estilo encabezados
    header_style = NamedStyle(name="header_style")
    header_style.font = Font(bold=True, color="FFFFFF", size=12)
    header_style.fill = PatternFill(start_color=COLOR_HEADER, fill_type="solid")
    header_style.alignment = Alignment(horizontal="center", vertical="center")
    header_style.border = Border(
        bottom=Side(style="thick"),
        left=Side(style="thin", color="FFFFFF"),
        right=Side(style="thin", color="FFFFFF")
    )

    # Estilo celdas normales
    cell_style = NamedStyle(name="cell_style")
    cell_style.font = Font(size=11)
    cell_style.alignment = Alignment(vertical="center")
    cell_style.border = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD")
    )

    # Estilo para proveedores importantes (ejemplo)
    destacado_style = NamedStyle(name="destacado_style")
    destacado_style.font = Font(bold=True, color="333333")
    destacado_style.fill = PatternFill(start_color=COLOR_DESTACADO, fill_type="solid")

    # Registrar estilos
    for style in [header_style, cell_style, destacado_style]:
        if style.name not in wb.named_styles:
            wb.add_named_style(style)

    # ========== ENCABEZADO ==========
    # Título principal
    ws.merge_cells('A1:F1')
    ws['A1'] = "REPORTE DE PROVEEDORES - CONTROL DE ACCESO"
    ws['A1'].font = Font(bold=True, size=14, color=COLOR_HEADER)
    ws['A1'].alignment = Alignment(horizontal="center")

    # Subtítulo
    ws.merge_cells('A2:F2')
    ws['A2'] = f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws['A2'].font = Font(italic=True, size=10, color="666666")
    ws['A2'].alignment = Alignment(horizontal="center")

    # Filtros aplicados
    filtros = []
    if proveedor_id:
        try:
            proveedor = Proveedor.objects.get(prov_id=proveedor_id)
            filtros.append(f"Proveedor: {proveedor.prov_nombre}")
        except Proveedor.DoesNotExist:
            pass
    if fecha_exacta: filtros.append(f"Fecha exacta: {fecha_exacta}")
    if fecha_inicio: filtros.append(f"Desde: {fecha_inicio}")
    if fecha_fin: filtros.append(f"Hasta: {fecha_fin}")
    
    ws.merge_cells('A3:F3')
    ws['A3'] = "Filtros: " + (" | ".join(filtros) if filtros else "Ningún filtro aplicado")
    ws['A3'].font = Font(size=9, color="555555")
    ws['A3'].alignment = Alignment(horizontal="center")

    # Encabezados de columnas
    headers = ["FECHA", "PROVEEDOR", "HORA ENTRADA", "HORA SALIDA", "ESTADO", "COMENTARIOS"]
    ws.append(headers)
    
    # Aplicar estilo a encabezados
    for col in range(1, len(headers)+1):
        cell = ws.cell(row=4, column=col)
        cell.style = header_style
        # Ajustar ancho de columnas
        if col == 2:  # Columna de Proveedor
            ws.column_dimensions[get_column_letter(col)].width = 30
        elif col == 6:  # Columna de Comentarios
            ws.column_dimensions[get_column_letter(col)].width = 40
        else:
            ws.column_dimensions[get_column_letter(col)].width = 20

    # ========== DATOS ==========
    for row_idx, entrada in enumerate(queryset, start=5):
        # Formatear datos
        comentario = entrada.e_comentario if entrada.e_comentario and entrada.e_comentario != "None" else "-"
        
        ws.append([
            entrada.e_fecha.strftime('%d/%m/%Y') if entrada.e_fecha else '-',
            str(entrada.e_prov) if entrada.e_prov else '-',
            entrada.e_entrada.strftime('%H:%M') if entrada.e_entrada else '-',
            entrada.e_salida.strftime('%H:%M') if entrada.e_salida else '-',
            str(entrada.e_ee) if entrada.e_ee else '-',
            comentario
        ])

        # Aplicar estilos base
        for col in range(1, len(headers)+1):
            cell = ws.cell(row=row_idx, column=col)
            cell.style = cell_style

        # Efecto zebra
        if row_idx % 2 == 0:
            for col in range(1, len(headers)+1):
                ws.cell(row=row_idx, column=col).fill = PatternFill(start_color=COLOR_FONDO, fill_type="solid")

        # Destacar proveedores importantes (ahora usando prov_nombre)
        if entrada.e_prov and "importante" in entrada.e_prov.prov_nombre.lower():
            for col in range(1, len(headers)+1):
                ws.cell(row=row_idx, column=col).style = destacado_style

    # ========== PIE DE PÁGINA ==========
    total_registros = len(queryset)
    ws.merge_cells(f'A{row_idx+1}:F{row_idx+1}')
    ws[f'A{row_idx+1}'] = f"Total proveedores registrados: {total_registros} | © {datetime.now().year} Departamento de Logística"
    ws[f'A{row_idx+1}'].font = Font(italic=True, size=10, color="666666")
    ws[f'A{row_idx+1}'].alignment = Alignment(horizontal="center")
    ws[f'A{row_idx+1}'].fill = PatternFill(start_color="F0F0F0", fill_type="solid")

    # Congelar encabezados
    ws.freeze_panes = "A5"

    # Autoajustar alturas de filas
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[4].height = 22

    # ========== GENERAR RESPUESTA ==========
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename="reporte_proveedores.xlsx"'}
    )
    wb.save(response)
    return response
    
# Vista de Proveedores Ingresados
@login_required
@user_passes_test(is_swap_tthh)
def cobradores(request):
    # Filtro base: entradas finalizadas con cobrador no nulo
    cobradores_qs = Entrada.objects.filter(
        e_ee__ee_estado="Finalizado✅", 
        e_cob__isnull=False
    ).select_related('e_cob', 'e_ee')  # Optimización

    # Obtener parámetros del GET
    cobrador_id = request.GET.get("cobrador")
    fecha_exacta = request.GET.get("fecha_exacta")
    fecha_inicio = request.GET.get("fecha_inicio")
    fecha_fin = request.GET.get("fecha_fin")

    # Filtro por cobrador
    if cobrador_id and cobrador_id != "":
        cobradores_qs = cobradores_qs.filter(e_cob__cob_id=cobrador_id)

    # Filtros de fecha
    if fecha_exacta:
        cobradores_qs = cobradores_qs.filter(e_fecha=fecha_exacta)
    if fecha_inicio and fecha_fin:
        cobradores_qs = cobradores_qs.filter(e_fecha__range=[fecha_inicio, fecha_fin])
    elif fecha_inicio:
        cobradores_qs = cobradores_qs.filter(e_fecha__gte=fecha_inicio)
    elif fecha_fin:
        cobradores_qs = cobradores_qs.filter(e_fecha__lte=fecha_fin)

    # Obtener lista de cobradores ACTIVOS para el dropdown
    lista_cobradores = Cobrador.objects.filter(
        cob_estado="activo"
    ).order_by('cob_nombre')

    return render(request, "cobradores.html", {
        'cobradores': cobradores_qs.order_by('-e_fecha'),
        'lista_cobradores': lista_cobradores,
    })
    
# Excel de Cobradores
@login_required
@user_passes_test(is_swap_tthh)
def exportar_excel_cobradores(request):
    queryset = Entrada.objects.filter(e_ee__ee_estado="Finalizado✅", e_cob__isnull=False)

    cobrador_id = request.GET.get("cobrador")
    fecha_exacta = request.GET.get("fecha_exacta")
    fecha_inicio = request.GET.get("fecha_inicio")
    fecha_fin = request.GET.get("fecha_fin")

    if cobrador_id:
        queryset = queryset.filter(e_cob_id=cobrador_id)
    if fecha_exacta:
        queryset = queryset.filter(e_fecha=fecha_exacta)
    if fecha_inicio and fecha_fin:
        queryset = queryset.filter(e_fecha__range=[fecha_inicio, fecha_fin])
    elif fecha_inicio:
        queryset = queryset.filter(e_fecha__gte=fecha_inicio)
    elif fecha_fin:
        queryset = queryset.filter(e_fecha__lte=fecha_fin)

    queryset = queryset.order_by('-e_fecha')

    # ========== CONFIGURACIÓN EXCEL ==========
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Registro Cobradores"

    # --- Configuración de estilos ---
    COLOR_HEADER = "2E5C8A"  # Azul profesional
    COLOR_FONDO = "F9F9F9"   # Gris claro
    COLOR_DESTACADO = "FFD700"  # Amarillo dorado para destacar

    # Estilo encabezados
    header_style = NamedStyle(name="header_style")
    header_style.font = Font(bold=True, color="FFFFFF", size=12)
    header_style.fill = PatternFill(start_color=COLOR_HEADER, fill_type="solid")
    header_style.alignment = Alignment(horizontal="center", vertical="center")
    header_style.border = Border(
        bottom=Side(style="thick"),
        left=Side(style="thin", color="FFFFFF"),
        right=Side(style="thin", color="FFFFFF")
    )

    # Estilo celdas normales
    cell_style = NamedStyle(name="cell_style")
    cell_style.font = Font(size=11)
    cell_style.alignment = Alignment(vertical="center")
    cell_style.border = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD")
    )

    # Estilo para cobradores importantes (ejemplo)
    destacado_style = NamedStyle(name="destacado_style")
    destacado_style.font = Font(bold=True, color="333333")
    destacado_style.fill = PatternFill(start_color=COLOR_DESTACADO, fill_type="solid")

    # Registrar estilos
    for style in [header_style, cell_style, destacado_style]:
        if style.name not in wb.named_styles:
            wb.add_named_style(style)

    # ========== ENCABEZADO ==========
    # Título principal
    ws.merge_cells('A1:F1')
    ws['A1'] = "REPORTE DE COBRADORES - CONTROL DE ACCESO"
    ws['A1'].font = Font(bold=True, size=14, color=COLOR_HEADER)
    ws['A1'].alignment = Alignment(horizontal="center")

    # Subtítulo
    ws.merge_cells('A2:F2')
    ws['A2'] = f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws['A2'].font = Font(italic=True, size=10, color="666666")
    ws['A2'].alignment = Alignment(horizontal="center")

    # Filtros aplicados
    filtros = []
    if cobrador_id:
        try:
            cobrador = Cobrador.objects.get(cob_id=cobrador_id)
            filtros.append(f"Cobrador: {cobrador}")
        except Cobrador.DoesNotExist:
            pass
    if fecha_exacta: filtros.append(f"Fecha exacta: {fecha_exacta}")
    if fecha_inicio: filtros.append(f"Desde: {fecha_inicio}")
    if fecha_fin: filtros.append(f"Hasta: {fecha_fin}")
    
    ws.merge_cells('A3:F3')
    ws['A3'] = "Filtros: " + (" | ".join(filtros) if filtros else "Ningún filtro aplicado")
    ws['A3'].font = Font(size=9, color="555555")
    ws['A3'].alignment = Alignment(horizontal="center")

    # Encabezados de columnas
    headers = ["FECHA", "COBRADOR", "HORA ENTRADA", "HORA SALIDA", "ESTADO", "COMENTARIOS"]
    ws.append(headers)
    
    # Aplicar estilo a encabezados
    for col in range(1, len(headers)+1):
        cell = ws.cell(row=4, column=col)
        cell.style = header_style
        # Ajustar ancho de columnas
        if col == 2:  # Columna de Cobrador
            ws.column_dimensions[get_column_letter(col)].width = 30
        elif col == 6:  # Columna de Comentarios
            ws.column_dimensions[get_column_letter(col)].width = 40
        else:
            ws.column_dimensions[get_column_letter(col)].width = 20

    # ========== DATOS ==========
    for row_idx, entrada in enumerate(queryset, start=5):
        # Formatear datos
        comentario = entrada.e_comentario if entrada.e_comentario and entrada.e_comentario != "None" else "-"
        
        ws.append([
            entrada.e_fecha.strftime('%d/%m/%Y') if entrada.e_fecha else '-',
            str(entrada.e_cob) if entrada.e_cob else '-',
            entrada.e_entrada.strftime('%H:%M') if entrada.e_entrada else '-',
            entrada.e_salida.strftime('%H:%M') if entrada.e_salida else '-',
            str(entrada.e_ee) if entrada.e_ee else '-',
            comentario
        ])

        # Aplicar estilos base
        for col in range(1, len(headers)+1):
            cell = ws.cell(row=row_idx, column=col)
            cell.style = cell_style

        # Efecto zebra
        if row_idx % 2 == 0:
            for col in range(1, len(headers)+1):
                ws.cell(row=row_idx, column=col).fill = PatternFill(start_color=COLOR_FONDO, fill_type="solid")

        # Destacar cobradores importantes (ejemplo)
        if entrada.e_cob and "importante" in str(entrada.e_cob).lower():
            for col in range(1, len(headers)+1):
                ws.cell(row=row_idx, column=col).style = destacado_style

    # ========== PIE DE PÁGINA ==========
    total_registros = len(queryset)
    ws.merge_cells(f'A{row_idx+1}:F{row_idx+1}')
    ws[f'A{row_idx+1}'] = f"Total cobradores registrados: {total_registros} | © {datetime.now().year} Departamento de Logística"
    ws[f'A{row_idx+1}'].font = Font(italic=True, size=10, color="666666")
    ws[f'A{row_idx+1}'].alignment = Alignment(horizontal="center")
    ws[f'A{row_idx+1}'].fill = PatternFill(start_color="F0F0F0", fill_type="solid")

    # Congelar encabezados
    ws.freeze_panes = "A5"

    # Autoajustar alturas de filas
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[4].height = 22

    # ========== GENERAR RESPUESTA ==========
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename="reporte_cobradores.xlsx"'}
    )
    wb.save(response)
    return response

# Vista de todos los Servicios Tecnicos ya checkeados
@login_required
@user_passes_test(is_swap_tthh)
def vistast(request):
    trabajos_realizados = Tecnicos.objects.filter(tec_ee__ee_estado="Finalizado✅")
    funcionario = request.GET.get("funcionario")
    fecha_exacta = request.GET.get("fecha_exacta")
    fecha_desde = request.GET.get("fecha_desde")
    fecha_hasta = request.GET.get("fecha_hasta")
    
    if funcionario:
        trabajos_realizados = trabajos_realizados.filter(tec_nombre__icontains=funcionario)

    if fecha_exacta:
        trabajos_realizados = trabajos_realizados.filter(tec_fecha=fecha_exacta)

    if fecha_desde:
        trabajos_realizados = trabajos_realizados.filter(tec_fecha__gte=fecha_desde)

    if fecha_hasta:
        trabajos_realizados = trabajos_realizados.filter(tec_fecha__lte=fecha_hasta)
        
    lista_tecnicos = (
        Tecnicos.objects.order_by("tec_nombre")
        .values_list("tec_nombre", flat=True)
        .distinct()
    )

    #lista_montos = Tecnicosmonto.objects.all()
    return render(request, "vistast.html",{
        'trabajos_realizados': trabajos_realizados,
        'lista_tecnicos': lista_tecnicos,
        #'lista_montos': lista_montos,
    })
    
# Excel de Servicio Tecnico
@login_required
@user_passes_test(is_swap_tthh)
def ex_excel(request):
    # Filtrar SOLO los finalizados por defecto (como en vistast)
    trabajos = Tecnicos.objects.filter(tec_ee__ee_estado="Finalizado✅")

    funcionario = request.GET.get("funcionario")
    fecha_exacta = request.GET.get("fecha_exacta")
    fecha_desde = request.GET.get("fecha_desde")
    fecha_hasta = request.GET.get("fecha_hasta")
    estado = request.GET.get("estado")

    if funcionario:
        trabajos = trabajos.filter(tec_nombre__icontains=funcionario)  # Usar 'trabajos' en lugar de 'trabajos_realizados'
    if fecha_exacta:
        trabajos = trabajos.filter(tec_fecha=fecha_exacta)
    if fecha_desde:
        trabajos = trabajos.filter(tec_fecha__gte=fecha_desde)
    if fecha_hasta:
        trabajos = trabajos.filter(tec_fecha__lte=fecha_hasta)
    if estado:
        trabajos = trabajos.filter(tec_ee__ee_estado=estado)

    # Crear libro Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Trabajos Realizados"

    # =============================================
    # PALETA DE COLORES (COHERENTE CON TU SASS)
    # =============================================
    COLOR_INDIGO = "4B0082"       # $accent-indigo
    COLOR_PINK = "FF1493"         # $accent-pink
    COLOR_GOLD = "FFD700"         # $accent-gold
    COLOR_PRIMARY = "4361EE"      # $primary-color
    COLOR_SECONDARY = "3F37C9"    # $secondary-color
    COLOR_ACCENT = "4CC9F0"       # $accent-color
    COLOR_LIGHT_BG = "F8F9FA"     # $light-bg
    COLOR_DARK_TEXT = "2B2D42"    # $dark-text
    COLOR_LIGHT_TEXT = "8D99AE"   # $light-text
    COLOR_WHITE = "FFFFFF"        # $white
    COLOR_TABLE_HOVER = "F0F5FF"  # Similar a $table-hover

    # =============================================
    # ESTILOS PERSONALIZADOS (COMO TU DISEÑO WEB)
    # =============================================
    
    # Estilo para el encabezado (como tu thead con gradiente)
    header_style = NamedStyle(name="header_style")
    header_style.font = Font(
        bold=True, 
        color=COLOR_WHITE, 
        size=12, 
        name='Segoe UI'  # Coherente con tu $font-primary
    )
    header_style.fill = GradientFill(
        stop=("000000", COLOR_INDIGO, "6A5ACD"),
        type="linear",
        degree=45  # Igual que tu linear-gradient(45deg, ...)
    )
    header_style.border = Border(
        bottom=Side(border_style="thin", color=COLOR_INDIGO),
        left=Side(border_style="thin", color=COLOR_INDIGO),
        right=Side(border_style="thin", color=COLOR_INDIGO),
        top=Side(border_style="thin", color=COLOR_INDIGO)
    )
    header_style.alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True
    )
    
    # Estilo para datos normales (como tu tbody tr)
    data_style = NamedStyle(name="data_style")
    data_style.font = Font(
        size=11, 
        name='Segoe UI', 
        color=COLOR_DARK_TEXT  # $dark-text
    )
    data_style.border = Border(
        bottom=Side(border_style="thin", color=COLOR_LIGHT_TEXT + "20"),  # 20% opacity
        left=Side(border_style="thin", color=COLOR_LIGHT_TEXT + "20"),
        right=Side(border_style="thin", color=COLOR_LIGHT_TEXT + "20"),
        top=Side(border_style="thin", color=COLOR_LIGHT_TEXT + "20")
    )
    data_style.alignment = Alignment(
        vertical="center",
        wrap_text=True
    )
    
    # Estilo para filas pares (como tu tr:nth-child(even))
    even_row_style = NamedStyle(name="even_row_style")
    even_row_style.fill = PatternFill(
        start_color=COLOR_WHITE,
        end_color=COLOR_WHITE,
        fill_type="solid"
    )
    
    # Estilo para filas impares (como tu tr:nth-child(odd))
    odd_row_style = NamedStyle(name="odd_row_style")
    odd_row_style.fill = PatternFill(
        start_color=COLOR_LIGHT_BG + "CC",  # 80% opacity
        end_color=COLOR_LIGHT_BG + "CC",
        fill_type="solid"
    )
    
    # Estilo para hover (simulado en Excel)
    hover_style = NamedStyle(name="hover_style")
    hover_style.fill = PatternFill(
        start_color=COLOR_TABLE_HOVER,
        end_color=COLOR_TABLE_HOVER,
        fill_type="solid"
    )
    
    # Estilo para montos (formato monetario)
    money_style = NamedStyle(name="money_style")
    money_style.number_format = '#,##0'
    money_style.font = Font(
        size=11, 
        name='Segoe UI', 
        color=COLOR_DARK_TEXT
    )
    money_style.alignment = Alignment(
        horizontal="right",
        vertical="center"
    )
    
    # Estilo para fechas (como en tu tabla web)
    date_style = NamedStyle(name="date_style")
    date_style.number_format = 'DD/MM/YYYY'
    date_style.font = Font(
        size=11, 
        name='Segoe UI', 
        color=COLOR_DARK_TEXT
    )
    date_style.alignment = Alignment(
        horizontal="center",
        vertical="center"
    )
    
    # Estilo para estado (destacado como en tu web)
    def get_estado_style(estado):
        estado = str(estado).lower()
        
        style = NamedStyle(name=f"estado_{estado[:10]}")
        style.alignment = Alignment(horizontal="center", vertical="center")
        style.font = Font(size=11, name='Segoe UI', bold=True)
        
        if 'finalizado' in estado:
            style.font.color = COLOR_WHITE
            style.fill = PatternFill(
                start_color=COLOR_PRIMARY,
                end_color=COLOR_PRIMARY,
                fill_type="solid"
            )
        elif 'pendiente' in estado:
            style.font.color = COLOR_DARK_TEXT
            style.fill = PatternFill(
                start_color=COLOR_GOLD,
                end_color=COLOR_GOLD,
                fill_type="solid"
            )
        elif 'cancelado' in estado or 'cancelar' in estado:
            style.font.color = COLOR_WHITE
            style.fill = PatternFill(
                start_color=COLOR_PINK,
                end_color=COLOR_PINK,
                fill_type="solid"
            )
        else:
            style.font.color = COLOR_DARK_TEXT
            style.fill = PatternFill(
                start_color=COLOR_LIGHT_BG,
                end_color=COLOR_LIGHT_BG,
                fill_type="solid"
            )
        
        return style
    
    # Añadir estilos al libro
    for style in [header_style, data_style, money_style, date_style, 
                 even_row_style, odd_row_style, hover_style]:
        wb.add_named_style(style)
    
    # =============================================
    # CONSTRUCCIÓN DEL EXCEL (COHERENTE CON TU HTML)
    # =============================================
    
    # Encabezados (idénticos a los de tu tabla web)
    headers = [
        "Funcionario", "Sitio", "Cliente", 
        "Descripción", "Fecha", "H. Inicio", 
        "H. Salida", "Estado", "Monto (Gs)"
    ]
    ws.append(headers)
    
    # Aplicar estilo a los encabezados
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.style = header_style
    
    # Filas de datos (como en tu template)
    for t in trabajos:
        row = [
            str(t.tec_nombre),
            str(t.tec_sitios),
            t.tec_cliente,
            t.tec_descripcion,
            t.tec_fecha,
            str(t.tec_hinicio),
            str(t.tec_hfinal),
            str(t.tec_ee),
            int(t.tec_tm.tm_monto) if t.tec_tm else 0,
        ]
        ws.append(row)
        
        # Aplicar estilos a la fila recién añadida
        current_row = ws.max_row
        
        # Estilo alterno para filas (como tu :nth-child)
        row_style = even_row_style if current_row % 2 == 0 else odd_row_style
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=current_row, column=col)
            
            # Estilo base
            cell.style = data_style
            
            # Aplicar estilo de fila (fondo alterno)
            cell.fill = row_style.fill
            
            # Estilo especial para la columna de monto
            if col == 9:
                cell.style = money_style
                cell.value = int(cell.value) if cell.value else 0
            
            # Estilo especial para la columna de fecha
            if col == 5:
                cell.style = date_style
            
            # Estilo especial para la columna de estado
            if col == 8:
                estado_style = get_estado_style(cell.value)
                cell.style = estado_style
    
    # =============================================
    # AJUSTES FINALES (PARA MEJOR EXPERIENCIA)
    # =============================================
    
    # Ajustar el ancho de columnas automáticamente
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = (max_length + 2) * 1.1
        ws.column_dimensions[column].width = adjusted_width
    
    # Ajustes especiales para columnas específicas
    ws.column_dimensions['D'].width = 40  # Descripción más ancha
    ws.column_dimensions['I'].width = 15  # Monto con ancho fijo
    
    # Congelar encabezados (como tu thead sticky)
    ws.freeze_panes = "A2"
    
    # Añadir filtros (como tu interfaz de búsqueda)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    
    # Añadir efecto hover simulado (cada 5 filas para visualización)
    for row in range(2, ws.max_row + 1, 5):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            if col != 8:  # No aplicar a estado
                cell.fill = hover_style.fill
    
    # Añadir título con estilo (como tu h2)
    ws.insert_rows(1)

    ws.merge_cells(f'A1:{get_column_letter(len(headers))}1')

    title_cell = ws['A1']
    title_cell.value = "Historial de Trabajos"
    title_cell.font = Font(
        bold=True, 
        size=14, 
        color=COLOR_INDIGO,
        name='Segoe UI'
    )
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Añadir borde decorativo inferior (como tu ::after)
    deco_cell = ws['A2']
    deco_cell.fill = GradientFill(
        stop=(COLOR_INDIGO, COLOR_PINK),
        type="linear",
        degree=90  # Como tu gradient(90deg, ...)
    )
    ws.row_dimensions[2].height = 3
    
    # =============================================
    # GENERAR RESPUESTA
    # =============================================
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=trabajos_realizados.xlsx'
    wb.save(response)
    
    return response


# Vista de Visitas
@login_required
@user_passes_test(is_swap_tthh)
def visitas(request):
    # Filtro base: entradas con visita no nula y finalizadas
    visitas_qs = Entrada.objects.filter(
        e_visita__isnull=False,
        e_ee__ee_estado="Finalizado✅"
    ).order_by('-e_fecha')

    # Obtener parámetros de filtrado
    fecha_exacta = request.GET.get('fecha_exacta')
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')

    # Aplicar filtros de fecha
    if fecha_exacta:
        visitas_qs = visitas_qs.filter(e_fecha=fecha_exacta)
    if fecha_desde and fecha_hasta:
        visitas_qs = visitas_qs.filter(e_fecha__range=[fecha_desde, fecha_hasta])
    elif fecha_desde:
        visitas_qs = visitas_qs.filter(e_fecha__gte=fecha_desde)
    elif fecha_hasta:
        visitas_qs = visitas_qs.filter(e_fecha__lte=fecha_hasta)

    return render(request, "visita.html", {
        'visitas': visitas_qs,
    })
    
    
# Excel de Visitas
@login_required
@user_passes_test(is_swap_tthh)
def exportar_visitas_excel(request):
    # Obtener los mismos filtros que en la vista visitas()
    visitas_qs = Entrada.objects.filter(
        e_visita__isnull=False,
        e_ee__ee_estado="Finalizado✅"
    ).order_by('-e_fecha')

    # Aplicar filtros de fecha
    fecha_exacta = request.GET.get('fecha_exacta')
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')

    if fecha_exacta:
        visitas_qs = visitas_qs.filter(e_fecha=fecha_exacta)
    if fecha_desde and fecha_hasta:
        visitas_qs = visitas_qs.filter(e_fecha__range=[fecha_desde, fecha_hasta])
    elif fecha_desde:
        visitas_qs = visitas_qs.filter(e_fecha__gte=fecha_desde)
    elif fecha_hasta:
        visitas_qs = visitas_qs.filter(e_fecha__lte=fecha_hasta)

    # Crear el libro de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Visitas"

    # Estilos (corregido el fill_type)
    header_fill = PatternFill(start_color="4B0082", end_color="6A5ACD", fill_type="solid")  # Cambiado a "solid"
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(left=Side(style='thin'), 
                   right=Side(style='thin'), 
                   top=Side(style='thin'), 
                   bottom=Side(style='thin'))

    # Encabezados
    headers = ['Fecha', 'Visita', 'Entrada', 'Salida', 'Estado', 'Comentario']
    ws.append(headers)
    
    # Aplicar estilos a los encabezados
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    # Datos
    for visita in visitas_qs:
        ws.append([
            visita.e_fecha.strftime('%d/%m/%Y') if visita.e_fecha else '-',
            visita.e_visita if visita.e_visita else '-',
            visita.e_entrada.strftime('%H:%M') if visita.e_entrada else '-',
            visita.e_salida.strftime('%H:%M') if visita.e_salida else '-',
            str(visita.e_ee) if visita.e_ee else '-',
            visita.e_comentario if visita.e_comentario and visita.e_comentario != "None" else '-'
        ])

    # Ajustar anchos de columnas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.1
        ws.column_dimensions[column].width = adjusted_width

    # Preparar la respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename=visitas_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
    wb.save(response)
    
    return response

# Control de Funcionarios
@login_required
@user_passes_test(is_swap_tthh)
def funcionarios(request):
    funcionarios = Funcionarios.objects.filter(fun_estado="activo")
    
    funcionario = request.GET.get("funcionario")
    fecha_exacta = request.GET.get("fecha_exacta")
    fecha_desde = request.GET.get("fecha_desde")
    fecha_hasta = request.GET.get("fecha_hasta")
    
    if funcionario:
            funcionarios = funcionarios.filter(fun_nombres_apellidos__icontains=funcionario)
            
    if fecha_exacta:
        funcionarios = funcionarios.filter(fun_entrada=fecha_exacta)

    if fecha_desde:
        funcionarios = funcionarios.filter(fun_entrada__gte=fecha_desde)

    if fecha_hasta:
        funcionarios = funcionarios.filter(fun_entrada__lte=fecha_hasta)
        
    lista_funcionarios = Funcionarios.objects.filter(fun_estado="activo")
        
    return render(request, "func.html",{
        'funcionarios':funcionarios,
        'lista_funcionarios':lista_funcionarios,
    })


# Excel de Nomina de Funcionarios
@login_required
@user_passes_test(is_swap_tthh)
def exportar_nomina(request):
    # Obtener los mismos filtros que en la vista principal
    funcionarios = Funcionarios.objects.filter(fun_estado="activo")
    
    funcionario = request.GET.get("funcionario")
    fecha_exacta = request.GET.get("fecha_exacta")
    fecha_desde = request.GET.get("fecha_desde")
    fecha_hasta = request.GET.get("fecha_hasta")
    
    if funcionario:
        funcionarios = funcionarios.filter(fun_nombres_apellidos__icontains=funcionario)
            
    if fecha_exacta:
        funcionarios = funcionarios.filter(fun_entrada=fecha_exacta)

    if fecha_desde:
        funcionarios = funcionarios.filter(fun_entrada__gte=fecha_desde)

    if fecha_hasta:
        funcionarios = funcionarios.filter(fun_entrada__lte=fecha_hasta)
    
    # Crear el libro de Excel
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="nomina_funcionarios.xlsx"'
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Nómina de Funcionarios"
    
    # Estilos personalizados que combinan con tu diseño
    header_font = Font(name='Poppins', bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4B0082", end_color="4B0082", fill_type="solid")  # Índigo
    cell_fill_even = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
    cell_fill_odd = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # Encabezados
    columns = [
        ("Cédula", 15),
        ("Nombre Completo", 30),
        ("Correo", 25),
        ("Sueldo", 15),
        ("Celular", 15),
        ("Departamento", 20),
        ("Entrada", 15)
    ]
    
    for col_num, (column_title, column_width) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_num, value=column_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_alignment
        ws.column_dimensions[get_column_letter(col_num)].width = column_width
    
    # Datos
    for row_num, funcionario in enumerate(funcionarios, 2):
        # Obtener el nombre del departamento usando el __str__ del modelo Departamentos
        departamento = str(funcionario.fun_dpto) if funcionario.fun_dpto else ""
        
        row = [
            str(funcionario.fun_ci),  # Convertir a string por si acaso
            funcionario.fun_nombres_apellidos,
            funcionario.fun_correo if funcionario.fun_correo else "",
            str(funcionario.fun_sueldo),  # Convertir a string para evitar problemas
            funcionario.fun_cel,
            departamento,  # Usamos el string del departamento
            funcionario.fun_entrada.strftime("%d/%m/%Y") if funcionario.fun_entrada else ""
        ]
        
        for col_num, cell_value in enumerate(row, 1):
            cell = ws.cell(row=row_num, column=col_num, value=cell_value)
            cell.border = border
            cell.alignment = Alignment(vertical='center')
            
            # Alternar colores de fila como en tu diseño
            if row_num % 2 == 0:
                cell.fill = cell_fill_even
            else:
                cell.fill = cell_fill_odd
    
    # Ajustar el formato de las columnas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width if adjusted_width > 5 else 5
    
    wb.save(response)
    return response

# Vista de los Telefonos Internos
@login_required
@user_passes_test(is_swap_tthh)
def telefonos(request):
    telefonos = Telefonosinternos.objects.all()
    
    funcionario = request.GET.get("funcionario")
    
    if funcionario:
            telefonos = Telefonosinternos.objects.filter(ti_fun__fun_nombres_apellidos__icontains=funcionario)
        
    lista_funcionarios = Funcionarios.objects.filter(fun_estado="activo")
    return render(request, "telef.html",{
        'lista_funcionarios':lista_funcionarios,
        'telefonos':telefonos,
    })


# Creacion a PDF de los telefonos internos
@login_required
@user_passes_test(is_swap_tthh)
def exportar_pdf(request):
    # Obtener los mismos filtros que en la vista principal
    telefonos = Telefonosinternos.objects.all()
    
    funcionario = request.GET.get("funcionario")
    
    if funcionario:
        telefonos = Telefonosinternos.objects.filter(ti_fun__fun_nombres_apellidos__icontains=funcionario)
    
    # Crear el PDF
    response = HttpResponse(content_type='application/pdf')
    filename = f"telefonos_internos_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Configuración del documento
    doc = SimpleDocTemplate(
        response,
        pagesize=letter,
        rightMargin=inch/2,
        leftMargin=inch/2,
        topMargin=inch/2,
        bottomMargin=inch/2
    )
    
    elements = []
    
    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'Title',
        parent=styles['Heading1'],
        fontSize=16,
        alignment=1,
        spaceAfter=20,
        textColor=colors.HexColor('#4b0082')
    )
    
    header_style = ParagraphStyle(
        'Header',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.white,
        alignment=1
    )
    
    # Título del documento
    title = Paragraph("NÓMINA DE TELÉFONOS INTERNOS", title_style)
    elements.append(title)
    
    # Información de la empresa
    company_info = [
        Paragraph("<b>EMPRESA:</b> SERTEC S.R.L.", styles['Normal']),
        Paragraph("<b>DEPARTAMENTO:</b> Talento Humano", styles['Normal']),
        Paragraph(f"<b>FECHA DE GENERACIÓN:</b> {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal']),
        Spacer(1, 20)
    ]
    elements.extend(company_info)
    
    # Datos de la tabla
    data = [['NÚMERO', 'FUNCIONARIO']]  # Encabezados
    
    for telefono in telefonos:
        data.append([
            telefono.ti_numero,
            str(telefono.ti_fun) if telefono.ti_fun else ""
        ])
    
    # Crear tabla
    table = Table(data)
    
    # Estilo de la tabla
    table.setStyle(TableStyle([
        # Encabezado
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4b0082')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        
        # Filas alternas
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 1, colors.lightgrey),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('LEFTPADDING', (0, 0), (-1, -1), 5),
        ('RIGHTPADDING', (0, 0), (-1, -1), 5),
        
        # Destacar filas pares
        ('BACKGROUND', (0, 2), (-1, -1), colors.HexColor('#f8f9fa')),
        ('BACKGROUND', (0, 1), (-1, 1), colors.white),
        ('BACKGROUND', (0, 3), (-1, 3), colors.white),
    ]))
    
    elements.append(table)
    
    # Pie de página
    footer = [
        Spacer(1, 20),
        Paragraph(
            "Este documento es confidencial y para uso exclusivo de la empresa",
            ParagraphStyle(
                name='Footer',
                fontSize=8,
                textColor=colors.grey,
                alignment=1
            )
        )
    ]
    elements.extend(footer)
    
    # Construir el PDF
    doc.build(elements)
    
    return response


# Vista de los Ex's Funcionarios
@login_required
@user_passes_test(is_swap_tthh)
def ex_funcionarios(request):
    
    ex_funcionarios = Funcionarios.objects.filter(fun_estado="inactivo")
    funcionario = request.GET.get("funcionario")
    fecha_exacta = request.GET.get("fecha_exacta")
    fecha_desde = request.GET.get("fecha_desde")
    fecha_hasta = request.GET.get("fecha_hasta")
    
    if funcionario:
        ex_funcionarios = ex_funcionarios.filter(fun_nombres_apellidos__icontains=funcionario)

    if fecha_exacta:
        ex_funcionarios = ex_funcionarios.filter(fun_salida=fecha_exacta)

    if fecha_desde:
        ex_funcionarios = ex_funcionarios.filter(fun_salida__gte=fecha_desde)

    if fecha_hasta:
        ex_funcionarios = ex_funcionarios.filter(fun_salida__lte=fecha_hasta)
        
    lista_exfuncionarios = (
        Funcionarios.objects.filter(fun_estado="inactivo")
    )
    
    return render(request, "ex_fun.html",{
        "ex_funcionarios":ex_funcionarios,
        "lista_exfuncionarios":lista_exfuncionarios,
    })




import calendar
from datetime import date, datetime, timedelta
from calendar import month_name
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Border, Side, Alignment,
    NamedStyle, numbers
)
from openpyxl.styles import GradientFill

from django.shortcuts import render, redirect
from django.contrib import messages
from django.db import transaction
from django.urls import reverse
from django.contrib.auth.decorators import login_required, user_passes_test
from django.utils import timezone
from datetime import timedelta,datetime
from django.db.models import Count
from django.http import HttpResponse
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.views.generic import ListView
from django.views.generic import TemplateView
from django.db.models import Count
from django.utils import timezone
from django.core.paginator import EmptyPage, PageNotAnInteger, Paginator

from swap_home.models import *
from swap_informatica.models import *
from swap_porteria.models import *
from swap_serviciotecnico.models import *
from swap_tecnico.models import *
from swap_tthh.models import *
from .forms import *
# Create your views here.

# --- Proteger de otros Usuarios Logeados ---
def is_swap_serviciotecnico(user):
    return user.groups.filter(name='ServicioTecnico').exists()

# menu principal servicio tecnico
@login_required
@user_passes_test(is_swap_serviciotecnico)
def serviciotecnico(request):
    tarea = Tareas.objects.filter(tarea_dpto__dpto_nombre="General")
    tarea_dia = Tareadia.objects.filter(td_dia="General")
    
    return render(request, "serviciotecnico.html",{
        'tarea': tarea,
        'dia': tarea_dia,
    })
    
    
    


@login_required
@user_passes_test(is_swap_serviciotecnico)
def editar_trabajos_bulk(request):
    """
    Guarda múltiples filas seleccionadas desde veregistros.html.
    Espera:
      - selected_ids: lista de tec_id marcados
      - Por cada id: estado_<id>, monto_<id>
    """
    if request.method != "POST":
        return redirect('veregistro')

    selected_ids = request.POST.getlist('selected_ids')
    if not selected_ids:
        messages.warning(request, "No seleccionaste ningún registro.")
        return redirect('veregistro')

    # Cargar catálogos para resolver ids a objetos
    estados = {str(e.ee_id): e for e in Estadoentrada.objects.all()}
    montos  = {str(m.tm_id): m for m in Tecnicosmonto.objects.all()}

    actualizado = 0
    with transaction.atomic():
        for tec_id in selected_ids:
            try:
                t = Tecnicos.objects.select_for_update().get(tec_id=tec_id)
            except Tecnicos.DoesNotExist:
                continue

            estado_id = request.POST.get(f"estado_{tec_id}")
            monto_id  = request.POST.get(f"monto_{tec_id}")

            if estado_id and estado_id in estados:
                t.tec_ee = estados[estado_id]
            if monto_id and monto_id in montos:
                t.tec_tm = montos[monto_id]

            t.save(update_fields=["tec_ee", "tec_tm"])
            actualizado += 1

    messages.success(request, f"Se actualizaron {actualizado} registro(s).")
    # Volver a la vista con filtros GET si los mandaste como hidden "next"
    next_url = request.POST.get("next") or reverse("veregistro")
    return redirect("veregistro")


# ver los registros ingresados por los trabajodores
from datetime import datetime

class VerRegistroListView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    template_name = "veregistros.html"
    context_object_name = "trabajos_realizados"
    paginate_by = 25

    def test_func(self):
        return is_swap_serviciotecnico(self.request.user)

    def _clean_date(self, value: str):
        value = (value or "").strip()
        if not value:
            return None
        # input type=date manda YYYY-MM-DD
        try:
            return datetime.strptime(value, "%Y-%m-%d").date()
        except ValueError:
            return None

    def get_queryset(self):
        qs = (
            Tecnicos.objects
            .select_related("tec_ee", "tec_tm")
            .filter(tec_ee__ee_estado="Pendiente⚠️")
        )

        funcionarios_sel = [n.strip() for n in self.request.GET.getlist("funcionario") if n.strip()]
        fecha_exacta = self._clean_date(self.request.GET.get("fecha_exacta"))
        fecha_desde  = self._clean_date(self.request.GET.get("fecha_desde"))
        fecha_hasta  = self._clean_date(self.request.GET.get("fecha_hasta"))

        if funcionarios_sel:
            # OR por nombres (IN)
            qs = qs.filter(tec_nombre__in=funcionarios_sel)

        # ✅ prioridad: fecha_exacta
        if fecha_exacta:
            qs = qs.filter(tec_fecha=fecha_exacta)
        else:
            if fecha_desde:
                qs = qs.filter(tec_fecha__gte=fecha_desde)
            if fecha_hasta:
                qs = qs.filter(tec_fecha__lte=fecha_hasta)

        return qs.order_by("-tec_fexacta", "-tec_hexacta", "-tec_id")

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)

        ctx["funcionarios_sel"] = [n.strip() for n in self.request.GET.getlist("funcionario") if n.strip()]
        ctx["lista_tecnicos"] = (
            Tecnicos.objects.order_by("tec_nombre")
            .values_list("tec_nombre", flat=True)
            .distinct()
        )
        ctx["lista_estados"] = Estadoentrada.objects.all()
        ctx["lista_montos"] = Tecnicosmonto.objects.all()

        q = self.request.GET.copy()
        q.pop("page", None)
        ctx["querystring_sin_page"] = q.urlencode()

        return ctx



from django.shortcuts import redirect, get_object_or_404
# se realiza cambios correspondientes (estados u monto)
@login_required
@user_passes_test(is_swap_serviciotecnico)
def editar_trabajo(request, tec_id):
    if request.method == 'POST':
        trabajo = get_object_or_404(Tecnicos, pk=tec_id)
        nuevo_estado_id = request.POST.get('estado')
        nuevo_monto_id = request.POST.get('monto')

        if nuevo_estado_id:
            trabajo.tec_ee_id = nuevo_estado_id
        if nuevo_monto_id:
            trabajo.tec_tm_id = nuevo_monto_id
        trabajo.save()

    return redirect('veregistro')

# se añade nuevo monto
@login_required
@user_passes_test(is_swap_serviciotecnico)
def addmonto(request):
    if request.method == 'POST':
        form = MontoForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('addmonto')  # Redirige a donde necesites
    else:
        form = MontoForm()
    return render(request,"addmonto.html",{
        'form':form,
    })


# se ve el historial de todos los trabajos
MESES_ES = {
    1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril", 5: "Mayo", 6: "Junio",
    7: "Julio", 8: "Agosto", 9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre"
}
class HistorialView(LoginRequiredMixin, UserPassesTestMixin, TemplateView):
    template_name = "historial.html"
    paginate_by = 50  # 25/50/100 según tu gusto

    def test_func(self):
        return is_swap_serviciotecnico(self.request.user)

    # -----------------------------
    # Helpers de parseo robusto
    # -----------------------------
    def _clean_str(self, v: str) -> str:
        return (v or "").strip()

    def _clean_int(self, v):
        v = self._clean_str(v)
        if not v:
            return None
        try:
            return int(v)
        except ValueError:
            return None

    def _clean_date(self, v):
        v = self._clean_str(v)
        if not v:
            return None
        try:
            # input type="date" => YYYY-MM-DD
            return datetime.strptime(v, "%Y-%m-%d").date()
        except ValueError:
            return None

    # -----------------------------
    # Construcción del gráfico
    # -----------------------------
    def _build_chart(self):
        filtro_tipo = self._clean_str(self.request.GET.get("filtro_tipo")) or "30_dias"

        # params posibles
        mes_inicio = self._clean_int(self.request.GET.get("mes_inicio"))
        mes_fin = self._clean_int(self.request.GET.get("mes_fin"))
        anio_inicio = self._clean_int(self.request.GET.get("año_inicio"))
        anio_fin = self._clean_int(self.request.GET.get("año_fin"))
        mes_exacto = self._clean_int(self.request.GET.get("mes_exacto"))
        anio_exacto = self._clean_int(self.request.GET.get("año_exacto"))

        # Solo finalizados para el gráfico
        datos_query = Tecnicos.objects.filter(tec_ee__ee_estado="Finalizado✅")

        # combos del gráfico
        años_disponibles = datos_query.dates("tec_fecha", "year").order_by("-tec_fecha__year")
        meses_disponibles = [{"num": i, "nombre": MESES_ES[i]} for i in range(1, 13)]

        labels, valores = [], []

        # ------- RANGO MESES (FIX REAL) -------
        if filtro_tipo == "rango_meses":
            # defaults seguros
            hoy = timezone.now().date()
            mi = mes_inicio or 1
            mf = mes_fin or 12
            ai = anio_inicio or hoy.year
            af = anio_fin or hoy.year

            # normaliza si usuario lo pone al revés (o cruza años)
            if (ai, mi) > (af, mf):
                ai, af = af, ai
                mi, mf = mf, mi

            start_date = date(ai, mi, 1)
            last_day = calendar.monthrange(af, mf)[1]
            end_date = date(af, mf, last_day)

            datos = (
                datos_query
                .filter(tec_fecha__gte=start_date, tec_fecha__lte=end_date)
                .values("tec_fecha__year", "tec_fecha__month")
                .annotate(total=Count("tec_id"))
                .order_by("tec_fecha__year", "tec_fecha__month")
            )

            labels = [f"{MESES_ES[d['tec_fecha__month']]} {d['tec_fecha__year']}" for d in datos]
            valores = [d["total"] for d in datos]

        # ------- MES EXACTO (por día) -------
        elif filtro_tipo == "mes_exacto":
            hoy = timezone.now().date()
            year = anio_exacto or hoy.year
            month = mes_exacto or hoy.month

            # cantidad de días del mes
            num_days = calendar.monthrange(year, month)[1]

            datos_db = (
                datos_query
                .filter(tec_fecha__year=year, tec_fecha__month=month)
                .values("tec_fecha__day")
                .annotate(total=Count("tec_id"))
                .order_by("tec_fecha__day")
            )
            datos_dict = {d["tec_fecha__day"]: d["total"] for d in datos_db}

            labels = [f"{day:02d}/{month:02d}/{year}" for day in range(1, num_days + 1)]
            valores = [datos_dict.get(day, 0) for day in range(1, num_days + 1)]

        # ------- RANGO AÑOS (por año) -------
        elif filtro_tipo == "rango_años":
            hoy = timezone.now().date()
            ai = anio_inicio or hoy.year
            af = anio_fin or hoy.year
            if ai > af:
                ai, af = af, ai

            datos = (
                datos_query
                .filter(tec_fecha__year__gte=ai, tec_fecha__year__lte=af)
                .values("tec_fecha__year")
                .annotate(total=Count("tec_id"))
                .order_by("tec_fecha__year")
            )
            labels = [str(d["tec_fecha__year"]) for d in datos]
            valores = [d["total"] for d in datos]

        # ------- 30 DÍAS (por día) -------
        else:
            fecha_limite = timezone.now().date() - timedelta(days=30)
            date_range = [fecha_limite + timedelta(days=x) for x in range(31)]

            datos_db = (
                datos_query
                .filter(tec_fecha__gte=fecha_limite)
                .values("tec_fecha")
                .annotate(total=Count("tec_id"))
                .order_by("tec_fecha")
            )
            datos_dict = {d["tec_fecha"]: d["total"] for d in datos_db}

            labels = [d.strftime("%d/%m/%Y") for d in date_range]
            valores = [datos_dict.get(d, 0) for d in date_range]

        return {
            "labels": labels,
            "datos": valores,
            "años_disponibles": años_disponibles,
            "meses_disponibles": meses_disponibles,
            "filtro_actual": filtro_tipo,
        }

    # -----------------------------
    # Tabla + filtros + paginación
    # -----------------------------
    def _build_table(self):
        qs = Tecnicos.objects.select_related("tec_ee", "tec_tm").all()

        funcionarios_sel = [self._clean_str(n) for n in self.request.GET.getlist("funcionario") if self._clean_str(n)]
        fecha_exacta = self._clean_date(self.request.GET.get("fecha_exacta"))
        fecha_desde = self._clean_date(self.request.GET.get("fecha_desde"))
        fecha_hasta = self._clean_date(self.request.GET.get("fecha_hasta"))
        estado = self._clean_str(self.request.GET.get("estado"))

        if funcionarios_sel:
            qs = qs.filter(tec_nombre__in=funcionarios_sel)

        # prioridad: exacta; si está, ignorar desde/hasta
        if fecha_exacta:
            qs = qs.filter(tec_fecha=fecha_exacta)
        else:
            if fecha_desde:
                qs = qs.filter(tec_fecha__gte=fecha_desde)
            if fecha_hasta:
                qs = qs.filter(tec_fecha__lte=fecha_hasta)

        if estado:
            qs = qs.filter(tec_ee__ee_estado=estado)

        qs = qs.order_by("-tec_fecha", "-tec_id")

        paginator = Paginator(qs, self.paginate_by)
        page_number = self._clean_str(self.request.GET.get("page")) or "1"

        try:
            page_obj = paginator.page(page_number)
        except PageNotAnInteger:
            page_obj = paginator.page(1)
        except EmptyPage:
            page_obj = paginator.page(paginator.num_pages)

        # combos
        lista_tecnicos = (
            Tecnicos.objects.order_by("tec_nombre")
            .values_list("tec_nombre", flat=True)
            .distinct()
        )
        lista_estados = Estadoentrada.objects.all()
        lista_montos = Tecnicosmonto.objects.all()

        q = self.request.GET.copy()
        q.pop("page", None)

        return {
            "trabajos_realizados": page_obj.object_list,
            "funcionarios_sel": funcionarios_sel,

            "lista_tecnicos": lista_tecnicos,
            "lista_estados": lista_estados,
            "lista_montos": lista_montos,

            "is_paginated": paginator.num_pages > 1,
            "page_obj": page_obj,
            "paginator": paginator,
            "querystring_sin_page": q.urlencode(),
        }

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx.update(self._build_chart())
        ctx.update(self._build_table())
        return ctx

# excel
@login_required
@user_passes_test(is_swap_serviciotecnico)
def e_e(request):
    # ==========================
    # Helpers robustos
    # ==========================
    def clean_str(v):
        return (v or "").strip()

    def parse_date(v):
        v = clean_str(v)
        if not v:
            return None
        try:
            return datetime.strptime(v, "%Y-%m-%d").date()
        except ValueError:
            return None

    # ==========================
    # Query base + filtros
    # (MISMOS que tu tabla)
    # ==========================
    trabajos = (
        Tecnicos.objects
        .select_related("tec_ee", "tec_tm")
        .all()
    )

    funcionarios_sel = [clean_str(n) for n in request.GET.getlist("funcionario") if clean_str(n)]
    fecha_exacta = parse_date(request.GET.get("fecha_exacta"))
    fecha_desde = parse_date(request.GET.get("fecha_desde"))
    fecha_hasta = parse_date(request.GET.get("fecha_hasta"))
    estado = clean_str(request.GET.get("estado"))

    if funcionarios_sel:
        trabajos = trabajos.filter(tec_nombre__in=funcionarios_sel)

    # prioridad: exacta > rango
    if fecha_exacta:
        trabajos = trabajos.filter(tec_fecha=fecha_exacta)
    else:
        if fecha_desde:
            trabajos = trabajos.filter(tec_fecha__gte=fecha_desde)
        if fecha_hasta:
            trabajos = trabajos.filter(tec_fecha__lte=fecha_hasta)

    if estado:
        trabajos = trabajos.filter(tec_ee__ee_estado=estado)

    trabajos = trabajos.order_by("tec_fecha", "tec_id")

    # ==========================
    # Crear libro Excel
    # ==========================
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Trabajos Realizados"

    # =============================================
    # PALETA
    # =============================================
    AZUL_OSCURO = "2E5C8A"
    AZUL_MEDIO = "4361EE"
    AZUL_CLARO = "4CC9F0"
    AZUL_GRADIENTE_INICIO = "3A0CA3"
    AZUL_GRADIENTE_FIN = "4895EF"
    GRIS_CLARO = "F5F7FA"
    GRIS_MEDIO = "E0E5EC"
    BLANCO = "FFFFFF"
    TEXTO_OSCURO = "2B2D42"

    # =============================================
    # Estilos (FIX robusto)
    # =============================================
    def _named_style_names():
        """
        openpyxl puede devolver strings o NamedStyle en wb.named_styles.
        Esto lo normaliza a un set de nombres.
        """
        names = set()
        for s in wb.named_styles:
            if isinstance(s, str):
                names.add(s)
            else:
                # NamedStyle u otro
                n = getattr(s, "name", None)
                if n:
                    names.add(n)
        return names

    def add_style_safe(style: NamedStyle):
        if style.name not in _named_style_names():
            wb.add_named_style(style)

    header_style = NamedStyle(name="header_style")
    header_style.font = Font(bold=True, color=BLANCO, size=12, name="Segoe UI")
    header_style.fill = GradientFill(
        stop=(AZUL_GRADIENTE_INICIO, AZUL_GRADIENTE_FIN),
        type="linear",
        degree=45
    )
    header_style.border = Border(
        bottom=Side(border_style="medium", color=AZUL_OSCURO),
        left=Side(border_style="thin", color=AZUL_MEDIO),
        right=Side(border_style="thin", color=AZUL_MEDIO),
        top=Side(border_style="medium", color=AZUL_OSCURO),
    )
    header_style.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    data_style = NamedStyle(name="data_style")
    data_style.font = Font(size=11, name="Segoe UI", color=TEXTO_OSCURO)
    data_style.border = Border(
        bottom=Side(border_style="thin", color=GRIS_MEDIO),
        left=Side(border_style="thin", color=GRIS_MEDIO),
        right=Side(border_style="thin", color=GRIS_MEDIO),
        top=Side(border_style="thin", color=GRIS_MEDIO),
    )
    data_style.alignment = Alignment(vertical="center", wrap_text=True)

    even_row_style = NamedStyle(name="even_row_style")
    even_row_style.fill = PatternFill(start_color=BLANCO, end_color=BLANCO, fill_type="solid")

    odd_row_style = NamedStyle(name="odd_row_style")
    odd_row_style.fill = PatternFill(start_color=GRIS_CLARO, end_color=GRIS_CLARO, fill_type="solid")

    money_style = NamedStyle(name="money_style")
    money_style.number_format = '#,##0" Gs"'
    money_style.font = Font(size=11, name="Segoe UI", color=TEXTO_OSCURO)
    money_style.alignment = Alignment(horizontal="right", vertical="center")
    money_style.border = data_style.border

    date_style = NamedStyle(name="date_style")
    date_style.number_format = "DD/MM/YYYY"
    date_style.font = Font(size=11, name="Segoe UI", color=TEXTO_OSCURO)
    date_style.alignment = Alignment(horizontal="center", vertical="center")
    date_style.border = data_style.border

    time_style = NamedStyle(name="time_style")
    time_style.number_format = "HH:MM"
    time_style.font = Font(size=11, name="Segoe UI", color=TEXTO_OSCURO)
    time_style.alignment = Alignment(horizontal="center", vertical="center")
    time_style.border = data_style.border

    for st in (header_style, data_style, even_row_style, odd_row_style, money_style, date_style, time_style):
        add_style_safe(st)

    # =============================================
    # Construcción del Excel
    # =============================================
    headers = [
        "Funcionario", "Sitio", "Cliente",
        "Descripción", "Fecha", "H. Inicio",
        "H. Salida", "Estado", "Monto (Gs)"
    ]

    ws.merge_cells("A1:I1")
    title_cell = ws["A1"]
    title_cell.value = "HISTORIAL DE TRABAJOS"
    title_cell.font = Font(bold=True, size=14, color=AZUL_OSCURO, name="Segoe UI")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill = PatternFill(start_color=GRIS_CLARO, end_color=GRIS_CLARO, fill_type="solid")
    ws.row_dimensions[1].height = 24

    filtros_texto = []
    if funcionarios_sel:
        filtros_texto.append("Funcionario(s): " + ", ".join(funcionarios_sel))
    if estado:
        filtros_texto.append(f"Estado: {estado}")
    if fecha_exacta:
        filtros_texto.append(f"Fecha exacta: {fecha_exacta.strftime('%d/%m/%Y')}")
    else:
        if fecha_desde:
            filtros_texto.append(f"Desde: {fecha_desde.strftime('%d/%m/%Y')}")
        if fecha_hasta:
            filtros_texto.append(f"Hasta: {fecha_hasta.strftime('%d/%m/%Y')}")

    start_row = 2
    if filtros_texto:
        ws.merge_cells("A2:I2")
        subtitle_cell = ws["A2"]
        subtitle_cell.value = "Filtros aplicados: " + " | ".join(filtros_texto)
        subtitle_cell.font = Font(size=10, italic=True, color=AZUL_MEDIO, name="Segoe UI")
        subtitle_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[2].height = 22
        start_row = 3

    # Encabezados
    ws.append(headers)
    header_row = start_row
    for col in range(1, len(headers) + 1):
        ws.cell(row=header_row, column=col).style = header_style

    # Datos (FULL, sin paginación)
    for t in trabajos.iterator(chunk_size=2000):
        try:
            monto = int(t.tec_tm.tm_monto) if t.tec_tm and t.tec_tm.tm_monto is not None else 0
        except Exception:
            monto = 0

        row = [
            clean_str(t.tec_nombre),
            clean_str(t.tec_sitios),
            clean_str(t.tec_cliente),
            clean_str(t.tec_descripcion),
            t.tec_fecha,
            t.tec_hinicio,
            t.tec_hfinal,
            str(t.tec_ee) if t.tec_ee else "",
            monto,
        ]
        ws.append(row)

        r = ws.max_row
        fill = even_row_style.fill if r % 2 == 0 else odd_row_style.fill

        for c in range(1, 10):
            cell = ws.cell(row=r, column=c)
            cell.style = data_style
            cell.fill = fill
            if c == 5:
                cell.style = date_style
            elif c in (6, 7):
                cell.style = time_style
            elif c == 9:
                cell.style = money_style

    # Ajustes finales
    widths = {"A": 25, "B": 20, "C": 25, "D": 45, "E": 12, "F": 10, "G": 10, "H": 18, "I": 15}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = f"A{header_row}:I{ws.max_row}"

    ws.merge_cells(f"A{ws.max_row + 1}:I{ws.max_row + 1}")
    decor_cell = ws[f"A{ws.max_row}"]
    decor_cell.fill = PatternFill(start_color=AZUL_CLARO, end_color=AZUL_CLARO, fill_type="solid")
    ws.row_dimensions[ws.max_row].height = 3

    filename = f'Historial_Trabajos_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response




